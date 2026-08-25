"""Loads Infoleap's own manual coding taxonomy for open-ended verbatim
questions (Key Buying Factors / Reasons for Rejection / Reasons for
Cancelling) from the 3 hidden reference sheets in the Masterfile workbook.

These sheets (`Supernet | Net | Sub-net | Codelist | Codes`) are the real
market-research netting scheme Infoleap's human coders used — confirmed
exact match against the live dashboard's scraped Acceptor numbers (9/11
Supernet categories, 2026-07-27).

CORRECTION (2026-07-27): an earlier investigation (2026-07-24) concluded no
respondent-level linkage to these codes existed anywhere in the Masterfile —
that was WRONG. `data_updated` has 3 columns literally named
`MQ2a+MQ2b_KBF` / `MQ3a+MQ3b_Rejecter` / `MQ3a+MQ3b_Booked and cancelled`
holding each respondent's own assigned codes as a concatenated 3-digit
string (e.g. "001020117176180" = codes 001, 020, 117...). `load_code_map()`
below decodes those against this sheet's `Codes` column — used by
`utils.data_engine.DataEngine.reasons_table()` for an exact, deterministic
reproduction, and by `utils/verbatim_intel.py`'s "Reproduce Live Site
Categories" section (formerly an LLM approximation over sampled verbatim
text — superseded by this exact method once the linkage was found)."""
import openpyxl
from utils.data_engine import MASTERFILE_PATH, MASTER_CONFIG_PATH

# Maps netting sheet names (as used in masterfile) to the equivalent sheet
# names in RE_MIS_Master.xlsx — lets load_code_map() prefer the editable
# master config while falling back to the original masterfile if needed.
_MASTER_SHEET_MAP = {
    "MQ2a+MQ2b_KBF": "netting_KBF",
    "MQ3a+MQ3b_Rejecter": "netting_Rejector",
    "MQ3a+MQ3b_Booked and cancelled": "netting_Cancelled",
}

SEGMENT_SHEETS = {
    "Acceptor": "MQ2a+MQ2b_KBF",
    "Rejector": "MQ3a+MQ3b_Rejecter",
    "Cancelled": "MQ3a+MQ3b_Booked and cancelled",
}


def sheet_for(segment, broad_prefix):
    """Rejector/Cancelled each offer 2 question pairs: an mq2*-prefixed
    positive pair ("why considered/liked RE" — same framing as Acceptor's
    KBF questions) and an mq3*-prefixed negative pair ("why rejected/
    cancelled" — that segment's own Reasons sheet). SEGMENT_SHEETS alone
    only has one sheet per segment, which is right for Acceptor (only ever
    asked mq2*) but wrong for the mq2* pair under Rejector/Cancelled — this
    routes by the ACTUAL question framing, not just the segment."""
    if broad_prefix.startswith("mq2"):
        return SEGMENT_SHEETS["Acceptor"]
    return SEGMENT_SHEETS[segment]


def _parse_netting_sheet(ws):
    """Shared parser for any netting worksheet (masterfile or master config).
    Returns {code_str: (supernet, net, subnet, codelist)} dict."""
    code_map = {}
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i == 0:
            continue  # header row
        if len(row) < 5:
            continue
        supernet, net, subnet, codelist, code = row[0], row[1], row[2], row[3], row[4]
        if not supernet or code is None:
            continue
        code_str = str(code).strip().zfill(3)
        s_net = net or supernet
        s_sub = subnet or s_net
        s_code = codelist or s_sub
        code_map[code_str] = (supernet, s_net, s_sub, s_code)
    return code_map


def load_code_map(masterfile_path, sheet_name):
    """{code (3-digit zero-padded string): (Supernet, Net, Sub-net, Codelist)} —
    tries RE_MIS_Master.xlsx netting sheets first (editable), falls back to
    original masterfile sheets. This lets taxonomy updates (new Supernets/Nets,
    label corrections) be made in RE_MIS_Master.xlsx without touching the raw data.

    Codes are zero-padded to 3 digits since the sheet's Codes column isn't
    always consistently formatted (sometimes "2", sometimes "002") but
    respondent-level strings always use fixed 3-digit chunks."""
    # Try master config first (editable taxonomy)
    master_sheet = _MASTER_SHEET_MAP.get(sheet_name)
    if master_sheet and MASTER_CONFIG_PATH.exists():
        try:
            wb_master = openpyxl.load_workbook(str(MASTER_CONFIG_PATH), read_only=True, data_only=True)
            if master_sheet in wb_master.sheetnames:
                code_map = _parse_netting_sheet(wb_master[master_sheet])
                wb_master.close()
                if code_map:
                    return code_map
            wb_master.close()
        except Exception:
            pass  # fall through to masterfile

    # Fall back to original masterfile
    wb = openpyxl.load_workbook(masterfile_path, read_only=True, data_only=True)
    if sheet_name not in wb.sheetnames:
        raise ValueError(
            f"Netting sheet {sheet_name!r} not found in {masterfile_path!r}. "
            f"Available sheets: {wb.sheetnames}. Check SEGMENT_SHEETS against "
            f"the current Masterfile — sheet names can drift across monthly drops."
        )
    code_map = _parse_netting_sheet(wb[sheet_name])
    wb.close()
    return code_map
