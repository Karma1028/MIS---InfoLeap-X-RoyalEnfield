"""Chart builders for the Digital Showroom. Plotly for the main distribution
bars; Altair for the compact month-over-month trend lines."""
import html as _html
import plotly.graph_objects as go
import altair as alt
import pandas as pd
import streamlit as st
from styles.theme import RE_RED, CHART_SEQUENCE, BORDER, MUTED, INFOLEAP_GREEN, INFOLEAP_ORANGE, INFOLEAP_BLUE
from utils.stat_engine import calculate_significance

BRAND_CONSIDERED_COLOR = "#1B8A8A"
REASONS_COLOR = "#D6742D"
ADD_REPL_COLOR = "#2E3192"
BRAND_OWNED_COLOR = "#662D91"

PLOTLY_CONFIG = {"displayModeBar": False, "staticPlot": False}

# Fixed width (px) of every data table's leftmost "Category" row-label
# column (_render_html_table). Used as the LEFT MARGIN of stacked_bar_chart
# too, so a chart's x-axis bars start at the exact same horizontal offset
# as the table's data columns underneath it — the two live in separate
# rendering engines (Plotly SVG vs HTML table) with no shared layout, so
# this constant is the only thing keeping them aligned. Auto-sized table
# columns (the old behavior) varied per table depending on label length,
# which is what caused the persistent bar-to-column misalignment.
CATEGORY_COL_WIDTH = 130

# Soothing mix of RE + Infoleap brand colors — per user feedback ("colouring
# of the charts are not at all visible... use lighter colours... subtle and
# soothing"). Each entry blends a brand hue toward white so it reads as a
# pastel tint rather than the saturated solid brand swatch (which is right
# for logos/buttons but too harsh stacked across a whole chart).
def _lighten(hex_color, amount=0.45):
    """Blends hex_color toward white by `amount` (0=no change, 1=white)."""
    hex_color = hex_color.lstrip('#')
    r, g, b = int(hex_color[0:2], 16), int(hex_color[2:4], 16), int(hex_color[4:6], 16)
    r = round(r + (255 - r) * amount)
    g = round(g + (255 - g) * amount)
    b = round(b + (255 - b) * amount)
    return f"#{r:02x}{g:02x}{b:02x}"


def _blend(hex1, hex2, t=0.5):
    """Blends two hex colors — t=0 is hex1, t=1 is hex2."""
    hex1, hex2 = hex1.lstrip('#'), hex2.lstrip('#')
    r = round(int(hex1[0:2], 16) * (1 - t) + int(hex2[0:2], 16) * t)
    g = round(int(hex1[2:4], 16) * (1 - t) + int(hex2[2:4], 16) * t)
    b = round(int(hex1[4:6], 16) * (1 - t) + int(hex2[4:6], 16) * t)
    return f"#{r:02x}{g:02x}{b:02x}"


SOOTHING_PALETTE = [
    _lighten("#C8102E", 0.45),  # RE red
    _lighten("#2E3192", 0.45),  # Infoleap blue
    _lighten("#F7941D", 0.40),  # Infoleap orange
    _lighten("#39B54A", 0.45),  # Infoleap green
    _lighten("#662D91", 0.45),  # Infoleap purple
    _lighten("#1B8A8A", 0.40),  # teal
    _lighten("#8C8C8C", 0.35),  # neutral grey
    _lighten("#D4A017", 0.40),  # gold
    _lighten("#5B7DB1", 0.40),  # steel blue
    _lighten("#A65A8C", 0.40),  # mauve
]

# Soft fill tints for significant bars/segments — deeper-but-still-muted
# for 95% confidence, barely-there for the 90% directional tier, so a
# significant slice reads as "highlighted" without turning the chart into
# a traffic light.
SIG_FILL_95_UP = "#8FCB9B"
SIG_FILL_90_UP = "#D6EEDA"
SIG_FILL_95_DOWN = "#E3A0A0"
SIG_FILL_90_DOWN = "#F6DCDC"


def _sig_fill(marker):
    """Fill-color override for a significant bar/segment, or None to keep
    the category's own base color."""
    return {
        '▲': SIG_FILL_95_UP, '△': SIG_FILL_90_UP,
        '▼': SIG_FILL_95_DOWN, '▽': SIG_FILL_90_DOWN,
    }.get(marker)


_HIGH_MARKERS = ('▲', '△')
_LOW_MARKERS = ('▼', '▽')


def filter_sig_markers(col_markers, direction):
    """Blanks out markers per the sidebar's High/Low direction filter —
    'High only' drops ▼/▽ (renders as plain, no red anywhere), 'Low only'
    drops ▲/△ (no green anywhere), 'Both' passes everything through
    unchanged. Applied once here (single choke point) so every chart type
    and its paired data table — which read the same marker values — stay
    in sync automatically.

    col_markers: {column_name: [marker_per_row, ...]} as produced by
    compare_to_baseline_by_column(), or None."""
    if not col_markers or direction == "Both":
        return col_markers
    drop = _LOW_MARKERS if direction == "High only" else _HIGH_MARKERS
    return {
        col: ['' if m in drop else m for m in markers]
        for col, markers in col_markers.items()
    }


def _pct_label(v, marker="", gap=None):
    """Shared chart-label text: '-' for a genuine zero (matches
    _render_html_table's existing 0->'-' convention below), else a whole-
    percent value, with an optional trailing significance marker. Every
    chart's inline bar/slice text should route through this instead of
    building its own f"{v:.0f}%" so a 0% cell never reads differently on
    the chart than it does in the data table right underneath it.

    gap: the percentage-point gap vs the baseline this marker was tested
    against (this value - baseline value) — per user request ("if one
    section is significant how much significant it is it should be
    shown"), a bare ▲ only said THAT something was significant, not BY
    HOW MUCH. Shown only when a marker is actually present (a gap on a
    non-significant category is noise, not signal)."""
    import math as _math
    base = "-" if (v == 0 or _math.isnan(v)) else f"{v:.0f}%"
    if not marker:
        return base
    return f"{base} {marker}".strip()


def _sig_border(marker):
    """Maps a compare_to_baseline() marker to a (line_color, line_width)
    border so significance is visible directly on the chart mark (bar
    border / donut wedge border), not just in the table cell — per user
    request to make 95%/90% significance 'visually understandable' beyond
    the table alone."""
    if marker in ('▲', '▼'):
        return (SIG_DEEP_GREEN if marker == '▲' else SIG_DEEP_RED), 3
    if marker in ('△', '▽'):
        return (SIG_LIGHT_GREEN if marker == '△' else SIG_LIGHT_RED), 2.5
    return "rgba(0,0,0,0)", 0


def distribution_bar(table_df, title, color=RE_RED, sig_markers=None):
    """Horizontal bar of the 'All' column for a distribution_table() result
    (numeric=True). Expects row 0 to be the Base row (n=), which is dropped
    from the plotted rows but always shown in the title — per
    MIS_Dashboard_Requirements.docx 5.5: 'Sample Size (n=) Display — Every
    chart and table must show the base size... non-negotiable.'

    sig_markers: optional list (one per category row) of significance
    markers from stat_engine.compare_to_baseline(), appended to the bar's
    data label (e.g. "47% ▲") AND drawn as a colored border on the bar
    itself (deep border = 95%, light border = 90%).
    """
    base_n = table_df.iloc[0]['All']
    rows = table_df.iloc[1:]
    values = rows['All'].astype(float)

    # Lighter, soothing fill — per user feedback ("colouring of the charts
    # are not at all visible... use lighter colours... subtle and
    # soothing"); the caller's accent (RE red / Infoleap blue/purple/teal)
    # still sets the hue, just blended toward a pastel instead of solid.
    base_fill = _lighten(color, 0.40)
    _SIG_LABEL = {'▲': '↑', '▼': '↓', '△': '↑', '▽': '↓'}
    labels = []
    fill_colors, border_colors, border_widths = [], [], []
    for i, v in enumerate(values):
        marker = sig_markers[i] if sig_markers and i < len(sig_markers) else ''
        lbl = _SIG_LABEL.get(marker, marker)
        labels.append(_pct_label(v, lbl))
        bc, bw = _sig_border(marker)
        border_colors.append(bc)
        border_widths.append(bw)
        fill_colors.append(_sig_fill(marker) or base_fill)

    fig = go.Figure(go.Bar(
        x=values, y=rows['Unnamed: 0'], orientation='h',
        marker=dict(color=fill_colors, line=dict(color=border_colors, width=border_widths), cornerradius=4),
        text=labels, textposition='outside',
        textfont=dict(size=12, color="#1A1A1A", family="Inter, Segoe UI, sans-serif"),
        hovertemplate="<b>%{y}</b><br>%{x:.1f}% of base<extra></extra>",
        cliponaxis=False,
    ))
    fig.update_layout(
        title=dict(text=f"{title}  <span style='font-size:11px;color:{MUTED}'>(n={base_n:,.0f})</span>",
                    font=dict(size=15, color="#1A1A1A", family="Oswald, Inter, sans-serif")),
        height=max(420, 200 + 40 * len(rows)) if len(rows) >= 5 else max(300, 45 * len(rows)),
        margin=dict(l=10, r=70, t=44, b=10),
        plot_bgcolor='rgba(0,0,0,0)', paper_bgcolor='rgba(0,0,0,0)',
        xaxis=dict(title=None, showgrid=False, zeroline=False, showticklabels=True,
                   tickfont=dict(size=11, color=MUTED), ticksuffix="%",
                   range=[0, max(float(values.max()) * 1.32, 10)]),
        yaxis=dict(autorange='reversed', showgrid=False, tickfont=dict(size=12.5, color="#2B2B2B", family="Inter, Segoe UI, sans-serif")),
        font=dict(color="#2B2B2B", family="Inter, Segoe UI, sans-serif"),
        bargap=0.36,
        showlegend=False,
    )
    return fig


# Distinct 10-color qualitative palette — demographics categories (Age
# bands, Education levels, etc) can run to 6-8 rows, and CHART_SEQUENCE's
# 6 brand colors started repeating/clashing. Per user feedback ("bit
# cluttered and hard to understand") on the stacked bars — uses
# SOOTHING_PALETTE (defined above) for the same reason the rest of this
# module now does: lighter, soothing tints instead of solid brand swatches.
HIGHLIGHT_COL_TINT = "#F2C14E"  # warm gold wash for the custom combined column


def stacked_bar_chart(table_df, title, color_seq=None, col_sig_markers=None, highlight_col=None, allow_all_sig=False):
    """Stacked column chart for Demographics — one stacked column per
    period (All + each selected month), segments = the category rows
    (Age bands, Education levels, etc). Replaces the old single-period
    bar/donut per user request ('in place of column charts they want to
    see stacked bar chart').

    highlight_col: the user's custom Year+Month combined comparison
    column (app.py sidebar) — every segment's fill in that one column
    gets washed toward a warm gold tint (distinct from both the category
    palette and the green/red significance fills) so it's visually
    findable for comparison against the regular per-month columns.
    Significance still takes priority where both apply.

    Decluttered per follow-up feedback ('cluttered and hard to
    understand'): no more inline %-value text on every slice (the exact
    numbers are already in the data table directly below — this was pure
    duplication and the main source of overlap); significant segments
    (vs the unfiltered Overview baseline, per-month) now get BOTH a
    thick colored border AND a bold marker glyph centered in the slice,
    so they're visible against the rest of the stack rather than relying
    on a thin border alone; legend moved to a scrollable right-side
    column instead of a wrapped top strip so categories stay one-per-line
    and readable instead of run-together.
    """
    base_n = table_df.iloc[0]['All']
    cat_rows = table_df.iloc[1:].reset_index(drop=True)
    cols = [c for c in table_df.columns if c != 'Unnamed: 0']
    seq = color_seq or SOOTHING_PALETTE

    # "All" is the summary column — rendered more saturated than month
    # columns so it reads as the primary reference point at a glance.
    all_col_color = _lighten  # identity alias; used below to de-saturate months
    fig = go.Figure()
    for i, (_, row) in enumerate(cat_rows.iterrows()):
        label = row['Unnamed: 0']
        ys = [float(row[c]) for c in cols]
        base_color = seq[i % len(seq)]
        # Month columns get an extra 20% lightening vs "All" so the All bar
        # visually pops as the summary without needing a separate indicator.
        month_color = _lighten(base_color, 0.22)
        fill_colors, border_colors, border_widths, texts, text_sizes = [], [], [], [], []
        for c in cols:
            is_all = (c == 'All')
            markers = (col_sig_markers or {}).get(c)
            marker = markers[i] if markers and i < len(markers) else ''
            bc, bw = _sig_border(marker)
            if not bw and highlight_col and c == highlight_col:
                bc, bw = "#B8860B", 2
            if not bw and is_all:
                bc, bw = base_color, 1.5  # thin border on All column to separate it
            border_colors.append(bc)
            border_widths.append(bw if bw else 0)
            # BUG (found 2026-07-28, user report: "significance value...
            # not correctly visible"): this branch never called
            # _sig_fill(marker) -- a significant segment only ever got the
            # colored BORDER above, never a highlighted fill. In a stacked
            # bar, a thin border on a slice sandwiched between other
            # colored slices is easy to miss entirely. Significance now
            # takes priority over the highlight_col tint / All-saturation
            # / month-lightening for the FILL too, not just the border.
            _sig_fc = _sig_fill(marker)
            if _sig_fc:
                # Blend significance tint with category color (60/40) so the
                # category is still identifiable while significance is visible.
                # Full override made all significant segments look the same green.
                fill_colors.append(_blend(base_color, _sig_fc, 0.45))
            elif highlight_col and c == highlight_col:
                fill_colors.append(_blend(base_color, HIGHLIGHT_COL_TINT, 0.55))
            elif is_all:
                fill_colors.append(base_color)   # full saturation on All
            else:
                fill_colors.append(month_color)  # lighter on month columns
            v = float(row[c])
            # Per user request: percentage shown throughout every section,
            # not just the All column above a declutter threshold — every
            # segment in every column (All and month) always gets a label.
            # NOT the gap magnitude here, deliberately -- tried it
            # (2026-07-28) and a longer "45% ▲ +12pp" label doesn't fit a
            # thin stacked-bar slice, Plotly shrinks/clips it into an
            # unreadable squeeze. The data table right below (more room
            # per cell) still shows the gap via col_sig_gaps -- that's
            # where "how much significant" is answered, not the bar text.
            texts.append(_pct_label(v, marker))
            text_sizes.append(11 if is_all else 10)
        fig.add_trace(go.Bar(
            name=label, x=cols, y=ys,
            marker=dict(color=fill_colors, line=dict(color=border_colors, width=border_widths)),
            text=texts, textposition='inside',
            textfont=dict(size=text_sizes, color="#1A1A1A", family="Inter, Segoe UI, sans-serif"),
            insidetextanchor='middle',
            hovertemplate=f"<b>{label}</b><br>%{{x}}: %{{y:.1f}}%<extra></extra>",
        ))
    # Columns aren't guaranteed to sum to 100% — multi-select questions
    # (e.g. "what CC did you own before" when a respondent can report more
    # than one prior bike) legitimately total over 100%. A hardcoded
    # y-axis range of ~105 clipped those columns' top slices clean off
    # (bug report: "some of them showing uneven column heights" —
    # Additional+Replaced onward, where multi-select totals run >100%).
    # Range now tracks the tallest actual stack instead of assuming 100%.
    col_totals = [sum(float(row[c]) for _, row in cat_rows.iterrows()) for c in cols] if len(cat_rows) else [0]
    max_total = max(col_totals) if col_totals else 100
    # Significance column-level annotations — visible flags ABOVE each month column
    # that has at least one significant category. Much clearer than tiny in-bar symbols.
    # Positive-only (per user request): markers are already filtered to
    # High/▲/△ only before reaching here (filter_sig_markers, app.py), so
    # only the "higher" branches are ever reachable.
    sig_annotations = []
    ann_y = max_total + 6
    for c in cols:
        if (c == 'All' and not allow_all_sig) or not col_sig_markers:
            continue
        markers_this_col = col_sig_markers.get(c, [])
        has_95_up = '▲' in markers_this_col
        has_90_up = '△' in markers_this_col
        if not (has_95_up or has_90_up):
            continue
        if has_95_up:
            sym, col_c = 'High ✓', SIG_DEEP_GREEN
        else:
            sym, col_c = '~High', SIG_LIGHT_GREEN
        sig_annotations.append(dict(
            x=c, y=ann_y, text=sym,
            showarrow=False, xref='x', yref='y',
            font=dict(size=12, color=col_c, family="Inter, sans-serif"),
            bgcolor='rgba(255,255,255,0.75)', borderpad=2,
            bordercolor=col_c, borderwidth=1,
        ))

    # Legend rendered EXTERNALLY (see stacked_bar_legend_html / its use in
    # render_chart_with_table) in a Streamlit column to the left of this
    # chart, not inside the Plotly figure.
    #
    # Left margin = CATEGORY_COL_WIDTH (not a small constant like 10px):
    # the data table rendered right below this chart has its own leftmost
    # "Category" row-label column at that exact fixed width (see
    # _render_html_table) — the table's DATA columns (All, Aug'25, ...)
    # only start after that label column, so the chart's plot area (where
    # the bars actually live) has to start at the same x-offset or the
    # bars will never sit above their matching table column no matter how
    # symmetric the chart's own margins look in isolation.
    # Per user report: charts with many categories (Reasons & Motivations
    # routinely has 10-19 rows) squeeze small-value slices too thin to
    # read. 32px/row is fine up to 4 categories; steeper from 5 up
    # (lowered from an initial 7-category threshold per follow-up
    # request), scoped to this chart only per explicit instruction.
    _n_cats = len(cat_rows)
    if _n_cats < 5:
        plot_height = max(350, 60 + 45 * _n_cats)
    else:
        plot_height = max(450, 250 + 40 * _n_cats)
    fig.update_layout(
        barmode='stack',
        title=dict(
            text=(
                f"{title}  <span style='font-size:11px;color:{MUTED}'>(n={base_n:,.0f})</span>"
                + (f"<span style='font-size:9px;color:#B0AAA3;'>"
                   f"  ·  High ✓ = 95% significantly higher  ~High = 90% likely higher vs rest</span>"
                   if sig_annotations else "")
            ),
            font=dict(size=15, color="#1A1A1A", family="Oswald, Inter, sans-serif"),
        ),
        height=plot_height,
        margin=dict(l=CATEGORY_COL_WIDTH, r=10, t=44, b=30),
        plot_bgcolor='rgba(0,0,0,0)', paper_bgcolor='rgba(0,0,0,0)',
        yaxis=dict(title=None, showgrid=False, zeroline=False, ticksuffix="%",
                   range=[0, max(ann_y + 6, max_total + 5, 112 if sig_annotations else 105)],
                   tickfont=dict(size=11, color=MUTED)),
        xaxis=dict(showgrid=False, tickfont=dict(size=11, color="#2B2B2B", family="Inter, Segoe UI, sans-serif"),
                   tickangle=-35,
                   tickmode="array", tickvals=cols,
                   # Same abbreviation the table header uses ("August'2025"
                   # -> "Aug'25") — per user request that the chart's x-axis
                   # labels match the table's column headers, not just the
                   # bars/columns themselves.
                   ticktext=["<b>All</b>" if c == "All" else (f"★ {c}" if c == highlight_col else _month_header(c)) for c in cols]),
        font=dict(color="#2B2B2B", family="Inter, Segoe UI, sans-serif"),
        showlegend=False,
        bargap=0.28,
        annotations=sig_annotations or [],
    )
    return fig


def stacked_bar_legend_html(table_df, color_seq=None):
    """External HTML legend for stacked_bar_chart — rendered in a narrow
    Streamlit column to the LEFT of the chart (render_chart_with_table),
    using the exact same per-row color assignment as the chart itself so
    swatches match. Kept out of the Plotly figure so the plot's bars stay
    flush-left and aligned with the data table's columns underneath."""
    cat_rows = table_df.iloc[1:].reset_index(drop=True)
    seq = color_seq or SOOTHING_PALETTE
    items = "".join(
        f"<div style='display:flex;align-items:flex-start;gap:6px;padding:3px 0;'>"
        f"<span style='width:10px;height:10px;border-radius:2px;background:{seq[i % len(seq)]};"
        f"flex-shrink:0;margin-top:3px;'></span>"
        f"<span style='font-size:11px;color:#2B2B2B;line-height:1.35;'>{row['Unnamed: 0']}</span>"
        f"</div>"
        for i, (_, row) in enumerate(cat_rows.iterrows())
    )
    return f"<div style='padding-top:38px;'>{items}</div>"


# donut_chart (and its donut_legend_rows/donut_height helpers) removed
# 2026-07-27 — per user request, every chart in the app now uses one
# consistent stacked-bar shape; nothing calls chart_type="donut" anymore.


SIG_DEEP_GREEN = "#1A7A3C"   # 95% confidence, base >= 30
SIG_LIGHT_GREEN = "#B7E4C0"  # 90% directional, base >= 30
SIG_DEEP_RED = "#9E2A2A"     # 95% confidence, significantly LOWER
SIG_LIGHT_RED = "#F0C2C2"    # 90% directional, significantly lower


def render_sig_legend(active_confidence=0.95):
    """Single shared legend body — sidebar only. All in-chart sig notation
    references this; nothing is repeated inline on charts.

    active_confidence: the currently-selected sidebar significance level
    (0.95 or 0.90) — stated at the top of the legend so the toggle and the
    legend can never disagree about what's currently active."""
    st.markdown(
        f"<div style='font-size:12px;color:#3A3732;line-height:1.9;'>"
        f"<div style='font-weight:600;margin-bottom:2px;'>Currently showing: significantly higher only</div>"
        f"<div style='color:#7A7670;font-size:11px;margin-bottom:8px;'>"
        f"Both the 95% and 90% confidence tiers are shown together, positive direction only — "
        f"a category never gets flagged for being significantly LOWER."
        f"</div>"
        f"<div style='font-weight:600;margin-bottom:4px;'>Bar/table cells</div>"
        f"<div><span style='display:inline-block;width:10px;height:10px;border-radius:2px;background:{SIG_DEEP_GREEN};margin-right:6px;'></span><b>High ✓</b> — confirmed higher (95%)</div>"
        f"<div><span style='display:inline-block;width:10px;height:10px;border-radius:2px;background:{SIG_LIGHT_GREEN};margin-right:6px;'></span><b>~High</b> — likely higher (90%)</div>"
        f"<div style='margin-top:8px;font-weight:600;margin-bottom:4px;'>Segment comparison bars (Overview)</div>"
        f"<div>↑ on a bar = that segment is <b>significantly higher</b> than the others</div>"
        f"<div style='margin-top:8px;font-weight:600;margin-bottom:4px;'>Model Comparison tables</div>"
        f"<div><b>X higher ✓</b> — Model X wins at 95% confidence</div>"
        f"<div><b>X higher ~</b> — Model X likely wins at 90%</div>"
        f"<div><b>Similar</b> — no clear difference detected</div>"
        f"<div><b>Above/Below avg ✓</b> — vs overall market average</div>"
        f"<div style='margin-top:8px;color:#7A7670;font-size:11px;'>Base &lt; 30: never tested. "
        f"Method: unpooled two-proportion Z-test on 0–100 scale.</div>"
        f"</div>",
        unsafe_allow_html=True,
    )


_FULL_MONTH_NAMES = {
    "January", "February", "March", "April", "May", "June", "July",
    "August", "September", "October", "November", "December",
}


def _month_header(m):
    """Abbreviates a real month column ("August'2025" -> "Aug'25"). Quarter-
    combined columns (e.g. "JAS'25") are already short — pass through
    unchanged rather than mangling them (year[2:] on a 2-digit year is empty,
    which would silently truncate the label)."""
    if "'" not in m:
        return m  # quarter label like "Q2 FY25-26" — pass through unchanged
    name, year = m.split("'", 1)
    if name not in _FULL_MONTH_NAMES:
        return m
    return f"{name[:3]}'{year[2:]}"


def _render_html_table(table_df, sig_markers=None, accent=RE_RED, col_sig_markers=None, rollup_labels=None, highlight_col=None, col_sig_gaps=None, hide_below=None, allow_all_sig=False):
    """Renders a compact bordered HTML table matching the live dashboard's
    report-table look, with cells colour-highlighted for significance (deep
    green = 95% confidence, light green = 90% directional; red shades for
    significantly LOWER) instead of plain arrow text — per user request,
    base>=30 is already enforced upstream in stat_engine.calculate_significance
    (n<30 -> no marker at all).

    sig_markers: markers for the 'All' column only (backward compatible).
    col_sig_markers: optional {column_name: [markers]} from
    stat_engine.compare_to_baseline_by_column() — per user request
    ('significance testing should run month by month for a sample size
    equal or over 30 respondant'), applies the SAME highlighting to each
    month column independently, not just the aggregate 'All' column.

    The Base row (n=) is highlighted with the segment's accent color, since
    the live dashboard always keeps the base row visually prominent —
    'every chart and table must show the base size... non-negotiable.'

    rollup_labels: optional set of brand-rollup row labels (e.g.
    {"Royal Enfield", "HERO", "BAJAJ"}) — per user request to match the live site's nested
    look, every OTHER non-Base row is indented and shown in a lighter
    weight as a "member" of whichever rollup precedes it."""
    cols = ["Unnamed: 0", "All"] + [c for c in table_df.columns if c not in ("Unnamed: 0", "All")]
    base_row_check = table_df.iloc[0] if len(table_df) > 0 else {}
    low_base_cols = set()
    for c in cols:
        if c not in ("Unnamed: 0", "All"):
            try:
                if float(base_row_check.get(c, 0)) < 50:
                    low_base_cols.add(c)
            except (ValueError, TypeError):
                pass
    # Long brand-wise tables (rollup + many member rows) get a sticky header
    # + scrollable body instead of pushing the whole page down, plus zebra
    # striping on member rows and a subtle top-border between brand groups
    # — per user request to keep a long table 'aesthetically pleasing and
    # quite clean looking even when it is long', not just functionally
    # complete.
    is_long = len(table_df) > 12
    # The custom Year+Month combined comparison column (app.py sidebar)
    # gets a warm gold wash on every cell in its column, header included
    # — so it reads as a distinct, comparable column at a glance rather
    # than just another month.
    HIGHLIGHT_BG = "#FBE9C6"
    HIGHLIGHT_HEADER_BG = "#F2C14E"

    def _header_html(c):
        if c == 'Unnamed: 0':
            return 'Category'
        if c == 'All':
            return 'All'
        if c == highlight_col:
            # Cap the header to a fixed width with ellipsis — a long
            # user-typed custom label (e.g. "My Custom Combined Months")
            # was stretching this one column much wider than every other
            # month column. Full name still available via title= tooltip.
            ec = _html.escape(c)
            return f"<span title='{ec}' style='display:inline-block;max-width:78px;overflow:hidden;text-overflow:ellipsis;white-space:nowrap;vertical-align:bottom;'>★ {ec}</span>"
        return _month_header(c)

    # Category column gets a FIXED width (CATEGORY_COL_WIDTH) instead of
    # auto-sizing to its longest label — auto-sizing meant the column's
    # width (and therefore where the data columns start) varied wildly
    # between tables (e.g. "18 to 25 Years" vs "First Time Buyer of 2W
    # (Family owns a 2W and not a primary user)"), which is why the
    # stacked_bar_chart above never lined up with its table: the chart's
    # left margin is a constant (see CATEGORY_COL_WIDTH below), so it can
    # only match a table whose row-label column is ALSO a constant width.
    header_cells = "".join(
        f"<th style='padding:10px 14px;text-align:center;font-weight:600;font-size:0.79rem;"
        f"letter-spacing:0.02em;"
        f"color:{'#1F3864' if c == highlight_col else ('#92400E' if c in low_base_cols else '#f8fafc')};"
        f"background:{HIGHLIGHT_HEADER_BG if c == highlight_col else ('#FEF3C7' if c in low_base_cols else '#1e293b')};"
        f"border-bottom:none;"
        + (f"width:{CATEGORY_COL_WIDTH}px;min-width:{CATEGORY_COL_WIDTH}px;max-width:{CATEGORY_COL_WIDTH}px;"
           "white-space:normal;word-break:break-word;"
           if c == "Unnamed: 0" else
           ("max-width:90px;white-space:nowrap;" if c == highlight_col else "white-space:nowrap;"))
        + "position:sticky;top:0;z-index:2;'>"
        f"{_header_html(c)}{' ⚠' if c in low_base_cols else ''}</th>"
        for c in cols
    )
    body_rows = []
    member_n = 0
    for i, row in table_df.iterrows():
        is_base = (i == 0)
        is_rollup = (not is_base) and rollup_labels is not None and str(row['Unnamed: 0']) in rollup_labels
        is_member = (not is_base) and rollup_labels is not None and not is_rollup
        # Skip data rows where All column rounds to 0 — no signal to show.
        if not is_base and 'All' in row:
            try:
                if float(row['All']) < 0.5:
                    continue
            except (ValueError, TypeError):
                pass
        if is_member:
            member_n += 1
        elif is_rollup:
            member_n = 0
        cells = []
        for c in cols:
            val = row[c]
            is_low_base = c in low_base_cols
            if c == "Unnamed: 0":
                txt = str(val)
                _cat_w = f"width:{CATEGORY_COL_WIDTH}px;min-width:{CATEGORY_COL_WIDTH}px;max-width:{CATEGORY_COL_WIDTH}px;white-space:normal;word-break:break-word;"
                if is_base:
                    style = f"padding:7px 10px;font-weight:800;color:{accent};text-align:left;" + _cat_w
                elif is_member:
                    txt = f"&nbsp;&nbsp;&nbsp;&nbsp;↳ {txt}"
                    style = "padding:5px 10px;color:#6A665F;font-size:12px;text-align:left;" + _cat_w
                elif rollup_labels is not None:
                    style = f"padding:7px 10px;font-weight:700;border-top:2px solid {BORDER};text-align:left;" + _cat_w
                else:
                    style = "padding:6px 10px;text-align:left;" + _cat_w
            else:
                try:
                    val = float(val)
                    import math as _math
                    if _math.isnan(val):
                        txt = "-"
                    elif is_low_base and is_base:
                        txt = f"{val:,.0f}"          # show base count even when low
                    elif is_low_base:
                        txt = "-"                    # suppress all data rows
                    elif is_base:
                        txt = f"{val:,.0f}"
                    elif val == 0:
                        txt = "-"
                    elif is_member:
                        txt = f"{val:.0f}%"
                    else:
                        txt = f"{val:.0f}%"
                except (ValueError, TypeError):
                    txt = "-" if (is_low_base and not is_base) else str(val)
                if is_low_base and is_base:
                    # highlight low-base count in amber to signal caution
                    style = "padding:7px 10px;text-align:center;font-weight:800;color:#92400E;background:#FEF3C7;border-radius:4px;"
                elif is_base:
                    style = f"padding:7px 10px;text-align:center;font-weight:800;color:{accent};"
                elif is_member:
                    style = "padding:5px 10px;text-align:center;font-size:12px;color:#6A665F;"
                elif rollup_labels is not None:
                    style = f"padding:7px 10px;text-align:center;font-weight:700;border-top:2px solid {BORDER};"
                else:
                    style = "padding:6px 10px;text-align:center;"
                # Visual-only suppression: hide values below threshold in brand-wise tables
                if not is_base and not is_low_base and hide_below is not None:
                    try:
                        if float(row[c]) < hide_below:
                            txt = "-"
                    except (ValueError, TypeError):
                        pass
                if not is_base and not is_low_base:
                    # Per user instruction: significance NEVER runs on the
                    # aggregate 'All' column, anywhere — only on individual
                    # month columns. col_sig_markers only ever carries month
                    # keys now (see compare_to_baseline_by_column call sites).
                    marker = None
                    if col_sig_markers and c in col_sig_markers and (c != "All" or allow_all_sig):
                        col_markers = col_sig_markers[c]
                        marker = col_markers[i - 1] if i - 1 < len(col_markers) else ''
                    # Suppress marker display when value rounds to "0%" or is a dash
                    _suppress = (txt in ("0%", "-"))
                    if not _suppress and marker in ('▲', '△', '▼', '▽'):
                        txt = f"{txt} {marker}"
                    if not _suppress and marker == '▲':
                        style += f"background:{SIG_DEEP_GREEN};color:white;font-weight:700;"
                    elif not _suppress and marker == '△':
                        style += f"background:{SIG_LIGHT_GREEN};color:#1A1A1A;font-weight:600;"
                    elif not _suppress and marker == '▼':
                        style += f"background:{SIG_DEEP_RED};color:white;font-weight:700;"
                    elif not _suppress and marker == '▽':
                        style += f"background:{SIG_LIGHT_RED};color:#1A1A1A;font-weight:600;"
                # Custom combined column tint — only when significance
                # didn't already color this cell (sig stays the priority
                # signal; the gold wash is just "this is the custom
                # comparison column" context, not a finding).
                if highlight_col and c == highlight_col:
                    if "background:" not in style:
                        style += f"background:{HIGHLIGHT_BG};"
                    style += "max-width:90px;"
            cells.append(f"<td style='{style}border-bottom:1px solid {BORDER};'>{txt}</td>")
        if is_base:
            bg = "background:#f1f5f9;"
        elif is_rollup:
            bg = "background:#FBF8F3;"
        elif is_member and member_n % 2 == 0:
            bg = "background:#f8fafc;"
        elif not is_rollup and not is_member and i % 2 == 0:
            bg = "background:#f8fafc;"
        else:
            bg = "background:#ffffff;"
        body_rows.append(f"<tr style='{bg}'>" + "".join(cells) + "</tr>")

    wrapper_style = (
        f"overflow-x:auto;overflow-y:auto;max-height:460px;border:1px solid #cbd5e1;border-radius:10px;margin-top:0.4rem;background:#ffffff;box-shadow:0 2px 8px rgba(0,0,0,0.06);"
        if is_long else
        f"overflow-x:auto;border:1px solid #cbd5e1;border-radius:10px;margin-top:0.4rem;background:#ffffff;box-shadow:0 2px 8px rgba(0,0,0,0.06);"
    )
    html = (
        f"<div style='{wrapper_style}'>"
        f"<table style='min-width:100%;width:max-content;border-collapse:collapse;font-size:0.82rem;font-family:-apple-system,BlinkMacSystemFont,\"Segoe UI\",Roboto,Helvetica,Arial,sans-serif;'>"
        f"<thead><tr>{header_cells}</tr></thead><tbody>{''.join(body_rows)}</tbody></table></div>"
    )
    if is_long:
        st.caption(f"{len(table_df) - 1} rows — scroll within the table to see all of them.")
    st.markdown(html, unsafe_allow_html=True)


# DEAD — safe to remove next cleanup (treemap_chart: no call sites in app.py or any utils as of 2026-08-10)
def treemap_chart(table_df, title):
    """Treemap for many-category ranked lists (brand-wise Brand Considered/
    Brand Owned/Additional+Replaced, 8+ rows after cap_rows) — per repeated
    'too much bar chart' feedback. Treemap is the MR-correct alternative
    here: it's a share-of-mentions composition across many categories,
    where area naturally draws the eye to the largest contributors, which
    a long bar list does less effectively."""
    base_n = table_df.iloc[0]['All']
    rows = table_df.iloc[1:]
    values = rows['All'].astype(float)
    fig = go.Figure(go.Treemap(
        labels=rows['Unnamed: 0'], values=values, parents=[""] * len(rows),
        marker=dict(colors=values, colorscale=[[0, "#F6D9D9"], [1, RE_RED]], line=dict(width=1, color="white")),
        text=[_pct_label(v) for v in values], textinfo="label+text",
        textfont=dict(size=13, family="Inter, Segoe UI, sans-serif"),
        hovertemplate="<b>%{label}</b><br>%{value:.1f}% of base<extra></extra>",
    ))
    fig.update_layout(
        title=dict(text=f"{title}  <span style='font-size:11px;color:{MUTED}'>(n={base_n:,.0f})</span>",
                    font=dict(size=15, color="#1A1A1A", family="Oswald, Inter, sans-serif")),
        height=380, margin=dict(l=4, r=4, t=44, b=4),
        paper_bgcolor='rgba(0,0,0,0)',
        font=dict(family="Inter, Segoe UI, sans-serif"),
    )
    return fig


_SEG_COLORS = {
    "Overview":  RE_RED,
    "Acceptors": INFOLEAP_GREEN,
    "Rejectors": RE_RED,
    "Cancelled": INFOLEAP_ORANGE,
}
_SEG_ORDER = ["Overview", "Acceptors", "Rejectors", "Cancelled"]


def segment_comparison_bar(seg_tables, title, current_seg="Overview"):
    """Grouped horizontal bar — one row per category, one bar per segment.
    Current segment's bars are full-saturation with % labels; others are
    muted so the current segment reads as the focus while the comparison
    is still visible. This is the PRIMARY chart on every demographic section
    — replaces the stacked-across-months bar which answered a secondary
    question (monthly trend) as if it were the primary one."""
    if not seg_tables:
        return None
    first_tbl = list(seg_tables.values())[0]
    categories = list(first_tbl.iloc[1:]['Unnamed: 0'])

    # On Overview page: current_seg="Overview" → we want ALL 3 segment bars
    # at full saturation + text (they are equally the focus — the comparison
    # IS the point). "Overview" trace itself is skipped on this mode (it
    # would be a 4th bar averaging out the other three — not useful).
    # On a specific segment page: only that segment's bars are highlighted.
    _overview_mode = (current_seg == "Overview") or (current_seg not in seg_tables)

    # Sort categories: on Overview mode sort by Acceptors %, otherwise by
    # current segment's % — highest category ranks top.
    _sort_key = "Acceptors" if _overview_mode else current_seg
    if _sort_key in seg_tables:
        cur_tbl = seg_tables[_sort_key]
        cat_vals = {}
        for _, row in cur_tbl.iloc[1:].iterrows():
            cat_vals[row['Unnamed: 0']] = float(row['All'])
        categories = sorted(categories, key=lambda c: cat_vals.get(c, 0), reverse=True)

    max_val = 0
    fig = go.Figure()
    for seg_name in _SEG_ORDER:
        if seg_name not in seg_tables:
            continue
        if _overview_mode and seg_name == "Overview":
            continue  # skip the blended "Overview" trace on the comparison view
        tbl = seg_tables[seg_name]
        base_n = float(tbl.iloc[0]['All'])
        cat_map = {row['Unnamed: 0']: float(row['All']) for _, row in tbl.iloc[1:].iterrows()}
        values = [cat_map.get(c, 0.0) for c in categories]
        max_val = max(max_val, max(values) if values else 0)
        is_current = _overview_mode or (seg_name == current_seg)
        color = _SEG_COLORS.get(seg_name, "#8C8C8C")
        fill = _lighten(color, 0.25 if is_current else 0.60)
        border = color if is_current else "rgba(0,0,0,0)"
        
        texts = []
        for v, cat in zip(values, categories):
            _marker = ""
            if seg_name == "Acceptors" and "Rejectors" in seg_tables:
                try:
                    rej_tbl = seg_tables["Rejectors"]
                    n_acc = base_n
                    n_rej = float(rej_tbl.iloc[0]['All'])
                    rej_cat_map = {row['Unnamed: 0']: float(row['All']) for _, row in rej_tbl.iloc[1:].iterrows()}
                    v_rej = rej_cat_map.get(cat, 0.0)
                    
                    if n_acc >= 30 and n_rej >= 30:
                        sig_res = calculate_significance(v / 100.0, n_acc, v_rej / 100.0, n_rej, confidence=0.95)
                        # Positive-only, per user request: only ever flag
                        # significantly HIGHER, never significantly lower —
                        # same rule applied everywhere else in the app.
                        if sig_res.get('is_significant') and v > v_rej:
                            _marker = "▲"
                except Exception:
                    pass
            # Per user request: percentage shown on every bar, current
            # segment and comparison segments alike (previously only the
            # "current"/highlighted segment's bars got a label at all).
            texts.append(_pct_label(v, _marker))

        fig.add_trace(go.Bar(
            name=f"{seg_name} ({int(base_n):,})",
            x=values, y=categories, orientation='h',
            marker=dict(color=fill, line=dict(color=border, width=2 if is_current else 0), cornerradius=3),
            text=texts, textposition='outside',
            textfont=dict(size=10, color="#1A1A1A", family="Inter, Segoe UI, sans-serif"),
            opacity=1.0 if is_current else 0.72,
            hovertemplate=f"<b>{seg_name}</b><br>%{{y}}: %{{x:.1f}}%<extra></extra>",
            cliponaxis=False,
        ))

    fig.update_layout(
        barmode='group',
        title=dict(
            text=f"{title}  <span style='font-size:11px;color:{MUTED}'>Segment comparison — All periods</span>",
            font=dict(size=15, color="#1A1A1A", family="Oswald, Inter, sans-serif"),
        ),
        height=max(450, 52 * len(categories) + 120) if len(categories) >= 5 else max(360, 52 * len(categories) + 100),
        margin=dict(l=10, r=10, t=80, b=20),
        plot_bgcolor='rgba(0,0,0,0)', paper_bgcolor='rgba(0,0,0,0)',
        xaxis=dict(title=None, showgrid=True, gridcolor="#F0EDE8", zeroline=False,
                   ticksuffix="%", tickfont=dict(size=11, color=MUTED), tickangle=0,
                   range=[0, min(max_val * 1.18, 105)]),
        yaxis=dict(autorange='reversed', showgrid=False,
                   tickfont=dict(size=12, color="#2B2B2B", family="Inter, Segoe UI, sans-serif")),
        legend=dict(orientation='h', yanchor='bottom', y=1.02, xanchor='center', x=0.5,
                    font=dict(size=11), traceorder='normal',
                    bgcolor='rgba(255,255,255,0)', borderwidth=0),
        bargap=0.28, bargroupgap=0.06,
        font=dict(color="#2B2B2B", family="Inter, Segoe UI, sans-serif"),
    )
    return fig


def render_chart_with_table(table_df, title, color=RE_RED, sig_markers=None, key=None, chart_type="bar", col_sig_markers=None, table_df_html=None, rollup_labels=None, highlight_col=None, table_in_expander=False, col_sig_gaps=None, allow_all_sig=False):
    """Renders the chart, then the underlying data table right below it
    (per user requirement: tabular data must accompany every chart, not
    just the chart alone, and match the live site's table layout).
    chart_type='stacked_bar' for the per-month stacked column chart, the
    default (any other value) for a single-column horizontal bar — per
    user request (2026-07-27), every chart in the app now uses one of
    these two consistent shapes; earlier donut/treemap/brand-rollup
    variants were removed from this dispatch.

    Per explicit user instruction ('significance running on the All column
    which should not happen... everywhere the all section is present') —
    significance NEVER applies to the aggregate 'All' column, anywhere, in
    any table. sig_markers is accepted for signature compatibility but is
    no longer used to color the chart/table — only col_sig_markers (month
    columns) drives any highlighting now.

    table_df_html: per user request ('brand wise data is not showing the
    full table') — when the CHART needs a capped/ranked subset (treemaps
    cap_rows() to ~8 rows + 'Other' so they stay readable), the data TABLE
    underneath should still show every row. Pass the full uncapped table
    here and table_df stays the (possibly capped) chart-only data; when
    omitted, both chart and table use table_df as before."""
    html_table = table_df_html if table_df_html is not None else table_df
    base_n = table_df.iloc[0]['All']
    all_zero = table_df.iloc[1:]['All'].astype(float).sum() == 0
    if base_n == 0 or all_zero:
        st.info(f"{title}: base n=0 for the current selection — this table doesn't apply here (e.g. Acceptors have no 'brand owned instead of RE'). Table shown below for completeness.")
        _render_html_table(html_table, accent=color, col_sig_markers=col_sig_markers, rollup_labels=rollup_labels, highlight_col=highlight_col, col_sig_gaps=col_sig_gaps, allow_all_sig=allow_all_sig)
        return
    # Per user request: significance legend lives ONLY in the sidebar now,
    # not repeated as a per-chart popover.
    # Per user request: one consistent chart shape across every section —
    # stacked bar everywhere (the earlier donut/treemap/brand-rollup/
    # brand-rollup-grouped chart types are gone; every call site in the
    # app now passes chart_type="stacked_bar"). treemap_chart() itself is
    # kept as a standalone function — Verbatim Intelligence's netting-
    # classification feature calls it directly, unrelated to this dispatch.
    if chart_type == "stacked_bar":
        fig = stacked_bar_chart(table_df, title, col_sig_markers=col_sig_markers, highlight_col=highlight_col, allow_all_sig=allow_all_sig)
    else:
        fig = distribution_bar(table_df, title, color=color)
    if chart_type == "stacked_bar":
        # Legend rendered in its own Streamlit column to the LEFT of the
        # chart. CRITICAL: the data table must render INSIDE the same
        # right-hand column as the chart, not full-width below both
        # columns — otherwise the chart's bars sit in a narrower (5/6)
        # container while the table's columns span the full width, and
        # the two never line up no matter what the chart's own margins
        # are. Keeping chart + table in the same column is what actually
        # keeps bars aligned with their table columns.
        _leg_col, _content_col = st.columns([1, 8])
        with _leg_col:
            st.markdown(stacked_bar_legend_html(table_df), unsafe_allow_html=True)
        with _content_col:
            st.plotly_chart(fig, use_container_width=True, config=PLOTLY_CONFIG, key=key)
            st.markdown("<div style='height:14px'></div>", unsafe_allow_html=True)
            if table_in_expander:
                with st.expander("📋 Data Table", expanded=False):
                    _render_html_table(html_table, accent=color, col_sig_markers=col_sig_markers, rollup_labels=rollup_labels, highlight_col=highlight_col, col_sig_gaps=col_sig_gaps, allow_all_sig=allow_all_sig)
            else:
                _render_html_table(html_table, accent=color, col_sig_markers=col_sig_markers, rollup_labels=rollup_labels, highlight_col=highlight_col, col_sig_gaps=col_sig_gaps, allow_all_sig=allow_all_sig)
        return
    else:
        st.plotly_chart(fig, use_container_width=True, config=PLOTLY_CONFIG, key=key)
    if table_in_expander:
        with st.expander("📋 Data Table", expanded=False):
            _render_html_table(html_table, accent=color, col_sig_markers=col_sig_markers, rollup_labels=rollup_labels, highlight_col=highlight_col, col_sig_gaps=col_sig_gaps, allow_all_sig=allow_all_sig)
    else:
        _render_html_table(html_table, accent=color, col_sig_markers=col_sig_markers, rollup_labels=rollup_labels, highlight_col=highlight_col, col_sig_gaps=col_sig_gaps, allow_all_sig=allow_all_sig)


# DEAD — safe to remove next cleanup (zone_heatmap: no call sites in app.py or any utils as of 2026-08-10)
def zone_heatmap(zone_matrix, title):
    """Category x Zone heatmap — % of base per category within each zone,
    ignoring the sidebar Zone filter so all 5 zones show side by side.
    Standard MR cross-tab visual for spotting regional skews at a glance.
    zone_matrix: {category: {zone: pct}}"""
    categories = list(zone_matrix.keys())
    zones = list(next(iter(zone_matrix.values())).keys()) if zone_matrix else []
    z = [[zone_matrix[cat][zone] for zone in zones] for cat in categories]

    # Readability fix: the old colorscale topped out at solid RE-red, which
    # made the fixed-dark text unreadable on the highest cells. Capped the
    # colorscale at a lighter rose tone so black text stays legible across
    # every cell, added a colorbar for scale context, and enlarged the grid.
    fig = go.Figure(go.Heatmap(
        z=z, x=zones, y=categories,
        colorscale=[[0, "#FCFBF9"], [0.5, "#F6D9D9"], [1, "#E8A0A8"]],
        zmin=0, zmax=max((max(row) for row in z), default=10),
        text=[[f"{v:.0f}%" for v in row] for row in z],
        texttemplate="%{text}", textfont=dict(size=13, color="#1A1A1A", family="Inter, Segoe UI, sans-serif"),
        hovertemplate="<b>%{y}</b> in <b>%{x}</b>: %{z:.0f}%<extra></extra>",
        showscale=True,
        colorbar=dict(title=dict(text="%", side="right"), thickness=12, len=0.8),
        xgap=3, ygap=3,
    ))
    fig.update_layout(
        title=dict(text=f"{title} — by Zone", font=dict(size=15, color="#1A1A1A", family="Oswald, Inter, sans-serif")),
        height=max(280, 48 * len(categories) + 90),
        margin=dict(l=140, r=20, t=60, b=10),
        plot_bgcolor='rgba(0,0,0,0)', paper_bgcolor='rgba(0,0,0,0)',
        font=dict(family="Inter, Segoe UI, sans-serif", size=13, color="#2B2B2B"),
        xaxis=dict(side='top', tickfont=dict(size=13)),
        yaxis=dict(autorange='reversed', tickfont=dict(size=13), automargin=True),
    )
    return fig


def segment_trend_chart(seg_dfs, month_cols):
    """Overlaid 3-segment monthly count trend (Acceptors/Rejectors/Cancelled).
    G3-B spec: melts per-month counts into long df, maps color to Segment.
    seg_dfs: dict {label: DataFrame} — each df has a 'month_label' column.
    month_cols: ordered list of month label strings."""
    _SEG_COLORS_TREND = {
        "Acceptors": "#39B54A",
        "Rejectors": "#C8102E",
        "Cancelled": "#F7941D",
    }
    rows = []
    for seg_label, seg_df in seg_dfs.items():
        if seg_label == "Overview" or seg_df is None or len(seg_df) == 0:
            continue
        for m in month_cols:
            cnt = int((seg_df['month_label'] == m).sum())
            rows.append({
                "Segment": seg_label,
                "Month": m.split("'")[0][:3] + "'" + m.split("'")[1][2:] if "'" in m else m,
                "Count": cnt,
            })
    if not rows:
        return None
    long_df = pd.DataFrame(rows)
    color_scale = alt.Scale(
        domain=list(_SEG_COLORS_TREND.keys()),
        range=list(_SEG_COLORS_TREND.values()),
    )
    chart = (
        alt.Chart(long_df)
        .mark_line(point=True, strokeWidth=2.5)
        .encode(
            x=alt.X("Month:N", sort=None, title=None),
            y=alt.Y("Count:Q", title="Respondents"),
            color=alt.Color("Segment:N", scale=color_scale, legend=alt.Legend(title=None)),
            tooltip=["Segment:N", "Month:N", "Count:Q"],
        )
        .properties(height=240)
        .configure_view(strokeWidth=0)
        .configure_axis(grid=True, gridColor="#F0EDE8")
    )
    return chart


def month_trend_chart(table_df, month_cols, category_filter=None):
    """Altair month-over-month line trend for selected categories. Expects
    row 0 to be the Base row, dropped here; numeric=True table."""
    rows = table_df.iloc[1:]
    if category_filter:
        rows = rows[rows['Unnamed: 0'].isin(category_filter)]

    long_rows = []
    for _, r in rows.iterrows():
        for m in month_cols:
            long_rows.append({"Category": r['Unnamed: 0'], "Month": m.split("'")[0][:3], "Value": float(r[m])})
    long_df = pd.DataFrame(long_rows)

    chart = alt.Chart(long_df).mark_line(point=True, strokeWidth=2.5).encode(
        x=alt.X('Month:N', sort=None, title=None),
        y=alt.Y('Value:Q', title='%'),
        color=alt.Color('Category:N', scale=alt.Scale(range=CHART_SEQUENCE), legend=alt.Legend(title=None)),
        tooltip=['Category:N', 'Month:N', 'Value:Q'],
    ).properties(height=260).configure_view(strokeWidth=0).configure_axis(grid=True, gridColor=BORDER)
    return chart


import textwrap


def render_collapsible_reasons_table(tree_data, title, color="#D6742D", key_suffix=""):
    """Renders a 3-level hierarchical, collapsible Supernet -> Net -> Subnet table for open-ended netting taxonomy questions.
    Row backgrounds are color-coded by level per user request:
    - Level 1 (Supernet): Orange tint (#FFEDD5)
    - Level 2 (Net): Yellow tint (#FEF9C3)
    - Level 3 (Subnet): Light Blue tint (#E0F2FE)
    Significance highlights:
    - 95% confidence (▲): Dark Green (#1A7A3C, white text)
    - 90% confidence (△): Light Green (#B7E4C0, dark green text)
    - NO red highlights (positive increases only).
    """
    if not tree_data or not tree_data.get('supernets'):
        st.info("No data available for this question / selection.")
        return

    col_bases = tree_data['col_bases']
    base_label = tree_data['base_label']
    supernets = tree_data['supernets']
    cols = list(tree_data['columns'])
    low_base_cols = {c for c in cols if c != 'All' and col_bases.get(c, 0) < 50}

    expand_all = st.checkbox("Expand All Categories", value=False, key=f"expand_all_{key_suffix}")

    def _th_cell(c):
        if c == 'Category':
            return "<th class='col-cat'>Category / Netting Factor</th>"
        label = c.split("'")[0][:3] + "'" + c.split("'")[1][2:] if "'" in c else c
        return f"<th class='col-val'>{label}</th>"

    header_html = "".join([_th_cell('Category')] + [_th_cell(c) for c in cols])

    def _base_cell(c):
        if c == 'Category':
            return f"<td class='col-cat'><strong>{_html.escape(base_label)}</strong></td>"
        if c in low_base_cols:
            return "<td class='col-val'><strong>-</strong></td>"
        n_val = col_bases.get(c, 0)
        return f"<td class='col-val'><strong>{n_val:,}</strong></td>"

    base_row_html = "".join([_base_cell('Category')] + [_base_cell(c) for c in cols])

    def _val_cells(item, row_bg, is_bold=False):
        cells = []
        pcts = item.get('pcts', {})
        markers = item.get('sig_markers', {})
        for c in cols:
            is_low_base = c in low_base_cols
            txt = "-" if is_low_base else str(pcts.get(c, '-'))
            marker = "" if is_low_base else markers.get(c, '')
            style = f"padding: 8px 12px; text-align: right; white-space: nowrap; border-bottom: 1px solid #cbd5e1; background: {row_bg};"
            if is_bold:
                style += " font-weight: 700;"
            if txt not in ("0%", "-") and marker == '▲':
                style += f" background: {SIG_DEEP_GREEN} !important; color: white !important; font-weight: 700;"
            elif txt not in ("0%", "-") and marker == '△':
                style += f" background: {SIG_LIGHT_GREEN} !important; color: #0F5132 !important; font-weight: 600;"
            cells.append(f"<td style='{style}'>{txt}</td>")
        return "".join(cells)

    tbodies = []
    dynamic_css_rules = []

    # Row background tints per user instructions:
    # Level 1 (Supernet): Orange tint (#FFEDD5, text #7C2D12)
    # Level 2 (Net): Yellow tint (#FEF9C3, text #713F12)
    # Level 3 (Subnet): Light Blue tint (#E0F2FE, text #0C4A6E)
    SNET_BG = "#FFEDD5"
    NET_BG = "#FEF9C3"
    SUBNET_BG = "#E0F2FE"

    for s_idx, snet in enumerate(supernets):
        chk_s_id = f"chk_s_{key_suffix}_{s_idx}"
        checked_attr = "checked" if expand_all else ""

        s_cat_cell = (
            f"<td class='col-cat' style='background:{SNET_BG}; color:#7C2D12; border-bottom:1px solid #cbd5e1;'>"
            f"<label for='{chk_s_id}' class='snet-label'>"
            f"<input type='checkbox' id='{chk_s_id}' class='snet-chk' {checked_attr} />"
            f"<span class='snet-arrow'></span>"
            f"<strong>{_html.escape(snet['name'])}</strong>"
            f"</label>"
            f"</td>"
        )
        snet_row = f"<tr class='snet-row'>{s_cat_cell}{_val_cells(snet, row_bg=SNET_BG, is_bold=True)}</tr>"

        level2_3_rows = []
        for n_idx, net in enumerate(snet.get('nets', [])):
            chk_n_id = f"chk_n_{key_suffix}_{s_idx}_{n_idx}"
            subnets = net.get('subnets', [])
            has_subnets = len(subnets) > 0

            if has_subnets:
                n_cat_cell = (
                    f"<td class='col-cat indent-net' style='background:{NET_BG}; color:#713F12; border-bottom:1px solid #cbd5e1;'>"
                    f"<label for='{chk_n_id}' class='net-label'>"
                    f"<input type='checkbox' id='{chk_n_id}' class='net-chk' {checked_attr} />"
                    f"<span class='net-arrow'></span>"
                    f"<strong>[{_html.escape(net['name'])}]</strong>"
                    f"</label>"
                    f"</td>"
                )
                dynamic_css_rules.append(
                    f".supernet-body:has(#{chk_s_id}:checked):has(#{chk_n_id}:checked) tr.sub-{key_suffix}-{s_idx}-{n_idx} {{ display: table-row !important; }}"
                )
            else:
                n_cat_cell = f"<td class='col-cat indent-net' style='background:{NET_BG}; color:#713F12; border-bottom:1px solid #cbd5e1; text-align: left !important;'><span class='net-prefix'>-</span>[{_html.escape(net['name'])}]</td>"

            net_row = f"<tr class='net-row'>{n_cat_cell}{_val_cells(net, row_bg=NET_BG, is_bold=has_subnets)}</tr>"
            level2_3_rows.append(net_row)

            for sub_idx, sub in enumerate(subnets):
                chk_sub_id = f"chk_sub_{key_suffix}_{s_idx}_{n_idx}_{sub_idx}"
                items = sub.get('items', [])
                has_items = len(items) > 0

                if has_items:
                    sub_cat_cell = (
                        f"<td class='col-cat indent-subnet' style='background:{SUBNET_BG}; color:#0C4A6E; border-bottom:1px solid #cbd5e1;'>"
                        f"<label for='{chk_sub_id}' class='sub-label'>"
                        f"<input type='checkbox' id='{chk_sub_id}' class='sub-chk' {checked_attr} />"
                        f"<span class='sub-arrow'></span>"
                        f"<strong>[{_html.escape(sub['name'])}]</strong>"
                        f"</label>"
                        f"</td>"
                    )
                    dynamic_css_rules.append(
                        f".supernet-body:has(#{chk_s_id}:checked):has(#{chk_n_id}:checked):has(#{chk_sub_id}:checked) tr.itm-{key_suffix}-{s_idx}-{n_idx}-{sub_idx} {{ display: table-row !important; }}"
                    )
                else:
                    sub_cat_cell = f"<td class='col-cat indent-subnet' style='background:{SUBNET_BG}; color:#0C4A6E; border-bottom:1px solid #cbd5e1; text-align: left !important;'><span class='sub-prefix'>-</span>[{_html.escape(sub['name'])}]</td>"

                sub_row = f"<tr class='subnet-row sub-{key_suffix}-{s_idx}-{n_idx}'>{sub_cat_cell}{_val_cells(sub, row_bg=SUBNET_BG, is_bold=has_items)}</tr>"
                level2_3_rows.append(sub_row)

                for itm in items:
                    itm_cat_cell = f"<td class='col-cat indent-item' style='background:#FAFAF9; color:#334155; border-bottom:1px solid #cbd5e1; padding-left: 64px; text-align: left !important;'><span class='item-prefix'>-</span>{_html.escape(itm['name'])}</td>"
                    itm_row = f"<tr class='item-row itm-{key_suffix}-{s_idx}-{n_idx}-{sub_idx}'>{itm_cat_cell}{_val_cells(itm, row_bg='#FAFAF9', is_bold=False)}</tr>"
                    level2_3_rows.append(itm_row)

        tbody_html = f"<tbody class='supernet-body'>{snet_row}{''.join(level2_3_rows)}</tbody>"
        tbodies.append(tbody_html)

    extra_css = "\n".join(dynamic_css_rules)
    css_styles = f"""
    <style>
    .netting-table-card {{
        font-family: -apple-system, BlinkMacSystemFont, "Segoe UI", Roboto, Helvetica, Arial, sans-serif;
        margin: 6px 0 16px 0;
    }}
    .netting-table-title {{
        font-size: 0.95rem;
        font-weight: 700;
        color: #1e293b;
        margin-bottom: 8px;
    }}
    .netting-scroll-wrap {{
        border: 1px solid #cbd5e1;
        border-radius: 8px;
        overflow-x: auto;
        background: #ffffff;
        box-shadow: 0 1px 3px rgba(0,0,0,0.04);
    }}
    .netting-table {{
        width: 100%;
        border-collapse: collapse;
        font-size: 0.82rem;
    }}
    .netting-table th {{
        background: #f8fafc;
        color: #334155;
        font-weight: 700;
        font-size: 0.8rem;
        padding: 9px 12px;
        border-bottom: 2px solid #cbd5e1;
        text-align: right;
        white-space: nowrap;
        position: sticky;
        top: 0;
        z-index: 2;
    }}
    .netting-table th.col-cat {{
        text-align: left;
        min-width: 130px;
        max-width: 160px;
    }}
    .netting-table td.col-cat {{
        text-align: left !important;
    }}
    .base-row td {{
        background: #f1f5f9;
        color: #1e293b;
        font-weight: 700;
        border-bottom: 2px solid #cbd5e1;
    }}
    .supernet-body {{
        border-bottom: 2px solid #cbd5e1;
    }}
    .snet-row {{
        font-weight: 700;
        color: #7C2D12;
    }}
    .snet-chk, .net-chk, .sub-chk {{
        display: none !important;
    }}
    .snet-label, .net-label, .sub-label {{
        cursor: pointer;
        display: inline-flex;
        align-items: center;
        width: 100%;
        user-select: none;
    }}
    .snet-arrow::before {{
        content: '▶';
        display: inline-block;
        width: 14px;
        font-size: 0.72rem;
        color: #C2410C;
        margin-right: 6px;
    }}
    .snet-chk:checked + .snet-arrow::before {{
        content: '▼';
    }}
    .net-arrow::before {{
        content: '▶';
        display: inline-block;
        width: 12px;
        font-size: 0.68rem;
        color: #854D0E;
        margin-right: 6px;
    }}
    .net-chk:checked + .net-arrow::before {{
        content: '▼';
    }}
    .supernet-body tr.net-row,
    .supernet-body tr.subnet-row,
    .supernet-body tr.item-row {{
        display: none;
    }}
    .supernet-body:has(.snet-chk:checked) tr.net-row {{
        display: table-row !important;
    }}
    {extra_css}
    .net-row {{
        color: #713F12;
        font-size: 0.81rem;
    }}
    .subnet-row {{
        color: #0C4A6E;
        font-size: 0.78rem;
    }}
    .indent-net {{
        padding-left: 24px !important;
    }}
    .indent-subnet {{
        padding-left: 44px !important;
    }}
    .sub-arrow::before {{
        content: '▶';
        display: inline-block;
        width: 10px;
        font-size: 0.62rem;
        color: #0369A1;
        margin-right: 6px;
    }}
    .sub-chk:checked + .sub-arrow::before {{
        content: '▼';
    }}
    .net-prefix, .sub-prefix {{
        color: #64748b;
        margin-right: 6px;
        font-weight: bold;
        text-align: left;
        display: inline-block;
    }}
    .item-prefix {{
        color: #94a3b8;
        margin-right: 6px;
        text-align: left;
        display: inline-block;
    }}
    </style>
    """

    html_code = f"""
    {css_styles}
    <div class="netting-table-card">
    <div class="netting-table-title">{_html.escape(title)}</div>
    <div class="netting-scroll-wrap">
    <table class="netting-table">
    <thead>
    <tr class="header-row">{header_html}</tr>
    <tr class="base-row">{base_row_html}</tr>
    </thead>
    {''.join(tbodies)}
    </table>
    </div>
    <div style="font-size:0.75rem;color:#64748b;margin-top:6px;padding-left:2px;">
    💡 <em>Click any Supernet or Net row to expand/collapse underlying factors.</em>
    </div>
    </div>
    """

    if hasattr(st, 'html'):
        st.html(html_code)
    else:
        st.markdown(html_code, unsafe_allow_html=True)


def render_collapsible_brand_table(table_df, title, color="#2E3192", rollup_labels=None,
                                    col_sig_markers=None, col_sig_gaps=None,
                                    highlight_col=None, accent=None, key_suffix=""):
    """Collapsible brand → model table matching the open-end supernet→net UI.
    Each brand row is a clickable header; its model rows expand/collapse via
    CSS :has() — same mechanics as render_collapsible_reasons_table.
    Values < 3% are displayed as '-' (visual only; raw data unchanged)."""
    import html as _h
    if table_df is None or len(table_df) < 2:
        st.info("No data available.")
        return

    rollup_set = set(rollup_labels) if rollup_labels else set()
    accent_col = accent or color

    cols_data = [c for c in table_df.columns if c not in ("Unnamed: 0",)]
    base_row_check = table_df.iloc[0] if len(table_df) > 0 else {}
    def _is_bwt_col_visible(c):
        if c == "All":
            return True
        try:
            return float(base_row_check.get(c, 0)) >= 50
        except (ValueError, TypeError):
            return True

    display_cols = ["All"] + [c for c in cols_data if c != "All"]
    low_base_bwt_cols = set()
    for c in display_cols:
        if c != "All":
            try:
                if float(base_row_check.get(c, 0)) < 50:
                    low_base_bwt_cols.add(c)
            except (ValueError, TypeError):
                pass

    SIG_DEEP_GREEN = "#1A7A3C"
    SIG_LIGHT_GREEN = "#B7E4C0"
    SIG_DEEP_RED = "#9B1C1C"
    SIG_LIGHT_RED = "#FECACA"
    BRAND_BG = "#EEF2FF"
    MODEL_BG = "#F8F9FF"
    BASE_BG = "#f1f5f9"
    HIGHLIGHT_BG = "#FBE9C6"
    HIGHLIGHT_HDR_BG = "#F2C14E"

    def _th(c):
        if c == "Category":
            return f"<th class='bwt-cat' style='background:#1e293b;color:#f8fafc;'>Category</th>"
        lbl = c
        if "'" in c:
            parts = c.split("'")
            lbl = parts[0][:3] + "'" + parts[1][2:] if len(parts) > 1 else c
        bg = HIGHLIGHT_HDR_BG if c == highlight_col else "#1e293b"
        fg = "#1F3864" if c == highlight_col else "#f8fafc"
        return f"<th class='bwt-val' style='background:{bg};color:{fg};'>{_h.escape(lbl)}</th>"

    header_html = _th("Category") + "".join(_th(c) for c in display_cols)

    # Base row
    base_row = table_df.iloc[0]
    def _base_td(c):
        if c == "Category":
            return f"<td class='bwt-cat' style='font-weight:800;color:{accent_col};text-align:left;'>{_h.escape(str(base_row['Unnamed: 0']))}</td>"
        if c in low_base_bwt_cols:
            return f"<td class='bwt-val' style='font-weight:800;color:{accent_col};'>-</td>"
        val = base_row.get(c, "")
        try:
            return f"<td class='bwt-val' style='font-weight:800;color:{accent_col};'>{int(float(val)):,}</td>"
        except Exception:
            return f"<td class='bwt-val'>{_h.escape(str(val))}</td>"
    base_html = "<tr class='bwt-base-row'>" + _base_td("Category") + "".join(_base_td(c) for c in display_cols) + "</tr>"

    # Build brand → [models] structure from sorted_tbl
    brands = []
    current_brand = None
    for i, row in table_df.iloc[1:].iterrows():
        lbl = str(row["Unnamed: 0"])
        if lbl in rollup_set:
            current_brand = {"label": lbl, "row": row, "row_idx": i, "models": []}
            brands.append(current_brand)
        elif current_brand is not None:
            current_brand["models"].append({"label": lbl, "row": row, "row_idx": i})

    # For sig markers: they are indexed by position in table_df.iloc[1:]
    orig_labels = list(table_df.iloc[1:]["Unnamed: 0"])

    def _get_marker(row_label, col):
        if not col_sig_markers or col not in col_sig_markers:
            return ""
        try:
            idx = orig_labels.index(row_label)
            return col_sig_markers[col][idx] if idx < len(col_sig_markers[col]) else ""
        except (ValueError, IndexError):
            return ""

    def _row_all_hidden(row):
        # Never hide model rows — always show all brands and models.
        return False

    def _val_cell(val_raw, marker, is_model=False, col=None):
        if col in low_base_bwt_cols:
            style = f"background:{HIGHLIGHT_BG};" if col == highlight_col else ""
            return f"<td class='bwt-val' style='{style}'>-</td>"
        try:
            val = float(val_raw)
        except (ValueError, TypeError):
            return f"<td class='bwt-val'>-</td>"
        # Show "-" if NaN or value rounds to 0%
        import math
        if math.isnan(val) or round(val) == 0:
            display = "-"
        elif is_model:
            display = f"{val:.0f}%"
        else:
            display = f"{val:.0f}%"

        # Significance styling
        style = ""
        suppress = (display == "-" or display == "0%")
        if not suppress and marker == "▲":
            style = f"background:{SIG_DEEP_GREEN};color:white;font-weight:700;"
            display = f"{display} ▲"
        elif not suppress and marker == "△":
            style = f"background:{SIG_LIGHT_GREEN};color:#0F5132;font-weight:600;"
            display = f"{display} △"
        elif not suppress and marker == "▼":
            style = f"background:{SIG_DEEP_RED};color:white;font-weight:700;"
            display = f"{display} ▼"
        elif not suppress and marker == "▽":
            style = f"background:{SIG_LIGHT_RED};color:#1A1A1A;font-weight:600;"
            display = f"{display} ▽"
        if col == highlight_col and "background:" not in style:
            style += f"background:{HIGHLIGHT_BG};"
        return f"<td class='bwt-val' style='{style}'>{display}</td>"

    tbodies_html = []
    for b_idx, brand in enumerate(brands):
        chk_id = f"bwt_chk_{key_suffix}_{b_idx}"
        b_row = brand["row"]
        b_lbl = brand["label"]

        brand_cat_cell = (
            f"<td class='bwt-cat bwt-brand-cat'>"
            f"<label for='{chk_id}' class='bwt-brand-label'>"
            f"<input type='checkbox' id='{chk_id}' class='bwt-chk' />"
            f"<span class='bwt-arrow'></span>"
            f"<strong>{_h.escape(b_lbl)}</strong>"
            f"</label></td>"
        )
        brand_val_cells = "".join(
            _val_cell(b_row.get(c, ""), _get_marker(b_lbl, c), is_model=False, col=c)
            for c in display_cols
        )
        brand_row_html = f"<tr class='bwt-brand-row'>{brand_cat_cell}{brand_val_cells}</tr>"

        model_rows_html = ""
        for m in brand["models"]:
            m_lbl = m["label"]
            m_row = m["row"]
            # Skip row if every data value would display as "-" (all < 3% or zero)
            if _row_all_hidden(m_row):
                continue
            model_cat_cell = (
                f"<td class='bwt-cat bwt-model-cat'>"
                f"<span class='bwt-model-arrow'>↳</span>{_h.escape(m_lbl)}"
                f"</td>"
            )
            model_val_cells = "".join(
                _val_cell(m_row.get(c, ""), _get_marker(m_lbl, c), is_model=True, col=c)
                for c in display_cols
            )
            model_rows_html += f"<tr class='bwt-model-row'>{model_cat_cell}{model_val_cells}</tr>"

        tbodies_html.append(
            f"<tbody class='bwt-brand-body' id='bwt_body_{key_suffix}_{b_idx}'>"
            f"{brand_row_html}{model_rows_html}"
            f"</tbody>"
        )

    css = f"""
    <style>
    .bwt-wrap-{key_suffix} {{
        overflow-x: auto;
        border: 1px solid #cbd5e1;
        border-radius: 10px;
        margin-top: 0.5rem;
        background: #ffffff;
        box-shadow: 0 2px 8px rgba(0,0,0,0.06);
    }}
    .bwt-wrap-{key_suffix} table {{
        width: 100%;
        border-collapse: collapse;
        font-size: 0.82rem;
        font-family: -apple-system, BlinkMacSystemFont, "Segoe UI", Roboto, sans-serif;
    }}
    .bwt-wrap-{key_suffix} th {{
        background: #1e293b;
        color: #f8fafc;
        font-weight: 700;
        font-size: 0.79rem;
        padding: 8px 8px;
        border-bottom: 2px solid #cbd5e1;
        white-space: nowrap;
        position: sticky;
        top: 0;
        z-index: 2;
    }}
    .bwt-wrap-{key_suffix} th.bwt-cat {{
        text-align: left;
        min-width: 120px;
        max-width: 135px;
        width: 130px;
    }}
    .bwt-wrap-{key_suffix} th.bwt-val {{
        text-align: center;
        min-width: 52px;
        width: 1%;
    }}
    .bwt-base-row td {{
        background: {BASE_BG};
        border-bottom: 2px solid #cbd5e1;
        padding: 7px 8px;
    }}
    .bwt-base-row td.bwt-val {{
        text-align: center;
        font-weight: 800;
    }}
    .bwt-brand-row td {{
        background: {BRAND_BG};
        padding: 8px 12px;
        border-top: 2px solid #c7d2fe;
        border-bottom: 1px solid #e0e7ff;
    }}
    .bwt-brand-cat {{
        text-align: left !important;
    }}
    .bwt-val {{
        text-align: center;
    }}
    .bwt-model-row td {{
        background: {MODEL_BG};
        padding: 6px 12px;
        border-bottom: 1px solid #e2e8f0;
    }}
    .bwt-model-cat {{
        text-align: left !important;
        color: #475569;
        font-size: 0.79rem;
        padding-left: 28px !important;
    }}
    .bwt-model-arrow {{
        color: #94a3b8;
        margin-right: 6px;
        font-weight: bold;
    }}
    .bwt-chk {{
        display: none !important;
    }}
    .bwt-brand-label {{
        cursor: pointer;
        display: inline-flex;
        align-items: center;
        width: 100%;
        user-select: none;
        color: #312e81;
        font-weight: 700;
    }}
    .bwt-arrow::before {{
        content: '▶';
        display: inline-block;
        width: 14px;
        font-size: 0.72rem;
        color: {color};
        margin-right: 6px;
        transition: transform 0.15s;
    }}
    .bwt-chk:checked + .bwt-arrow::before {{
        content: '▼';
    }}
    .bwt-brand-body .bwt-model-row {{
        display: none;
    }}
    .bwt-brand-body:has(.bwt-chk:checked) .bwt-model-row {{
        display: table-row !important;
    }}
    </style>
    """

    html_out = f"""
    {css}
    <div class="bwt-wrap-{key_suffix}">
    <table>
    <thead>
      <tr>{header_html}</tr>
      {base_html}
    </thead>
    {''.join(tbodies_html)}
    </table>
    </div>
    <div style="font-size:0.75rem;color:#64748b;margin-top:6px;padding-left:2px;">
    💡 <em>Click any brand row to expand/collapse its models.</em>
    </div>
    """

    if hasattr(st, 'html'):
        st.html(html_out)
    else:
        st.markdown(html_out, unsafe_allow_html=True)


def open_end_tree_to_excel(tree_data, title="Open End", show_sig=True):
    """Flatten reasons tree_data to formatted Excel bytes for download."""
    import io
    from openpyxl import Workbook
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

    if not tree_data or not tree_data.get("supernets"):
        return None

    cols = tree_data["columns"]
    col_bases = tree_data["col_bases"]
    base_label = tree_data.get("base_label", "Base")
    supernets = tree_data["supernets"]

    wb = Workbook()
    ws = wb.active
    ws.title = title[:31]

    # openpyxl requires 8-char ARGB hex (FF prefix = fully opaque)
    HDR_FILL  = PatternFill(patternType="solid", fgColor="FFC8102E")
    SNET_FILL = PatternFill(patternType="solid", fgColor="FFFFEDD5")
    NET_FILL  = PatternFill(patternType="solid", fgColor="FFFEF9C3")
    SUB_FILL  = PatternFill(patternType="solid", fgColor="FFE0F2FE")
    BASE_FILL = PatternFill(patternType="solid", fgColor="FFF3F4F6")

    thin = Side(style="thin", color="FFCBD5E1")
    bdr = Border(left=thin, right=thin, top=thin, bottom=thin)

    def _clean(val):
        if isinstance(val, float):
            s = "-" if val == 0 else f"{round(val)}%"
        else:
            s = str(val) if val is not None else "-"
        if not show_sig:
            s = s.replace(" ▲", "").replace(" △", "").replace("▲", "").replace("△", "").strip()
        return s

    # Header
    header = ["Level", "Category"] + cols
    for ci, h in enumerate(header, 1):
        cell = ws.cell(row=1, column=ci, value=h)
        cell.fill = HDR_FILL
        cell.font = Font(bold=True, color="FFFFFFFF", size=10)
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
        cell.border = bdr

    # Base row
    row_idx = 2
    for ci, v in enumerate(["Base", base_label] + [col_bases.get(c, 0) for c in cols], 1):
        cell = ws.cell(row=row_idx, column=ci, value=v)
        cell.fill = BASE_FILL
        cell.font = Font(bold=True, size=10)
        cell.alignment = Alignment(horizontal="left" if ci <= 2 else "center")
        cell.border = bdr
    row_idx += 1

    for snet in supernets:
        for ci, v in enumerate(["L1", snet["name"]] + [_clean(snet["pcts"].get(c, "-")) for c in cols], 1):
            cell = ws.cell(row=row_idx, column=ci, value=v)
            cell.fill = SNET_FILL
            cell.font = Font(bold=True, size=10, color="FF7C2D12")
            cell.alignment = Alignment(horizontal="left" if ci <= 2 else "center")
            cell.border = bdr
        row_idx += 1

        for net in snet.get("nets", []):
            for ci, v in enumerate(["L2", "  " + net["name"]] + [_clean(net["pcts"].get(c, "-")) for c in cols], 1):
                cell = ws.cell(row=row_idx, column=ci, value=v)
                cell.fill = NET_FILL
                cell.font = Font(size=10, color="FF713F12")
                cell.alignment = Alignment(horizontal="left" if ci <= 2 else "center")
                cell.border = bdr
            row_idx += 1

            for subnet in net.get("subnets", []):
                for ci, v in enumerate(["L3", "    " + subnet["name"]] + [_clean(subnet["pcts"].get(c, "-")) for c in cols], 1):
                    cell = ws.cell(row=row_idx, column=ci, value=v)
                    cell.fill = SUB_FILL
                    cell.font = Font(size=9, color="FF0C4A6E")
                    cell.alignment = Alignment(horizontal="left" if ci <= 2 else "center")
                    cell.border = bdr
                row_idx += 1

    ws.column_dimensions["A"].width = 6
    ws.column_dimensions["B"].width = 42
    for ci in range(3, len(header) + 1):
        ws.column_dimensions[ws.cell(row=1, column=ci).column_letter].width = 11
    ws.freeze_panes = "C2"

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()




