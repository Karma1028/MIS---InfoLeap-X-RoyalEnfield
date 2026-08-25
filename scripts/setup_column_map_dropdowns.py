"""
setup_column_map_dropdowns.py
Smart column-mapping reconciler for RE_MIS_Master.xlsx.

Run after pasting new raw data:
    python scripts/setup_column_map_dropdowns.py

Logic:
  1. Load current column_mapping (manual assignments).
  2. Load raw_data actual headers.
  3. For each mapped row:
       - raw_column found in raw_data  -> status: confirmed (no change)
       - raw_column NOT found          -> try fuzzy auto-match
           * match found               -> status: auto-matched (updates raw_column)
           * no match                  -> status: NEEDS REVIEW (blanks raw_column)
  4. Write status back to column_mapping sheet (new 'status' column).
  5. Refresh _raw_headers hidden sheet + DataValidation dropdown on col A.

User workflow:
  - Open column_mapping, filter status='NEEDS REVIEW', pick correct header from dropdown.
  - Re-run script to re-confirm, or just reload dashboard.
"""
import sys
import os
import difflib

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

import pandas as pd
import openpyxl
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.styles import PatternFill, Font

MASTER_PATH = "data/RE_MIS_Master.xlsx"
RAW_SHEET = "raw_data"
MAP_SHEET = "column_mapping"
HEADERS_SHEET = "_raw_headers"

FILL_OK    = PatternFill("solid", fgColor="D6F5D6")   # green  - confirmed
FILL_AUTO  = PatternFill("solid", fgColor="FFF3CD")   # amber  - auto-matched
FILL_NEEDS = PatternFill("solid", fgColor="F8D7DA")   # red    - needs review
FILL_WILD  = PatternFill("solid", fgColor="E8F4FD")   # blue   - wildcard (skip)


def get_raw_headers(path):
    probe = pd.read_excel(path, sheet_name=RAW_SHEET, nrows=0, header=0)
    header_row = 1 if all(str(c).startswith("Unnamed") for c in probe.columns[:5]) else 0
    df = pd.read_excel(path, sheet_name=RAW_SHEET, nrows=0, header=header_row)
    return df.columns.tolist()


def normalize(s):
    return str(s).lower().replace(" ", "_").replace("-", "_").strip()


def fuzzy_match(target, candidates, cutoff=0.75):
    """Return best match from candidates for target, or None if below cutoff."""
    norm_target = normalize(target)
    norm_map = {normalize(c): c for c in candidates}
    # 1. exact normalized match
    if norm_target in norm_map:
        return norm_map[norm_target]
    # 2. difflib close match
    matches = difflib.get_close_matches(norm_target, norm_map.keys(), n=1, cutoff=cutoff)
    if matches:
        return norm_map[matches[0]]
    # 3. substring: target inside candidate or vice versa
    for norm_c, orig_c in norm_map.items():
        if norm_target in norm_c or norm_c in norm_target:
            return orig_c
    return None


def reconcile():
    print(f"Loading {MASTER_PATH} ...")
    raw_headers = get_raw_headers(MASTER_PATH)
    raw_set = set(raw_headers)
    print(f"  raw_data: {len(raw_headers)} columns")

    map_df = pd.read_excel(MASTER_PATH, sheet_name=MAP_SHEET)
    print(f"  column_mapping: {len(map_df)} rows")

    results = []
    confirmed = auto_matched = needs_review = wildcards = 0

    for _, row in map_df.iterrows():
        raw_col = str(row.get("raw_column", "")).strip()
        internal = str(row.get("internal_name", "")).strip()
        desc = row.get("description", "")
        req = row.get("required", "")
        notes = row.get("notes", "")

        # Wildcard patterns (e.g. dq2a_*, aq5a_*) — skip reconciliation
        if "*" in raw_col:
            results.append({
                "raw_column": raw_col,
                "internal_name": internal,
                "description": desc,
                "required": req,
                "notes": notes,
                "status": "wildcard",
                "_fill": FILL_WILD,
            })
            wildcards += 1
            continue

        if raw_col in raw_set:
            # Confirmed: exact match
            results.append({
                "raw_column": raw_col,
                "internal_name": internal,
                "description": desc,
                "required": req,
                "notes": notes,
                "status": "confirmed",
                "_fill": FILL_OK,
            })
            confirmed += 1
        else:
            # Try auto-match
            match = fuzzy_match(raw_col, raw_headers)
            if match:
                print(f"  AUTO-MATCH: '{raw_col}' -> '{match}' (internal: {internal})")
                results.append({
                    "raw_column": match,
                    "internal_name": internal,
                    "description": desc,
                    "required": req,
                    "notes": f"[auto-matched from '{raw_col}'] {notes}".strip(),
                    "status": "auto-matched",
                    "_fill": FILL_AUTO,
                })
                auto_matched += 1
            else:
                print(f"  NEEDS REVIEW: '{raw_col}' not found in raw_data (internal: {internal})")
                results.append({
                    "raw_column": "",   # blank -> user picks from dropdown
                    "internal_name": internal,
                    "description": desc,
                    "required": req,
                    "notes": f"[was: '{raw_col}'] {notes}".strip(),
                    "status": "NEEDS REVIEW",
                    "_fill": FILL_NEEDS,
                })
                needs_review += 1

    print(f"\nReconciliation summary:")
    print(f"  Confirmed   : {confirmed}")
    print(f"  Auto-matched: {auto_matched}")
    print(f"  Needs review: {needs_review}")
    print(f"  Wildcards   : {wildcards}")

    # --- Write back to Excel ------------------------------------------------
    wb = openpyxl.load_workbook(MASTER_PATH)

    # Refresh _raw_headers hidden sheet
    if HEADERS_SHEET in wb.sheetnames:
        del wb[HEADERS_SHEET]
    ws_h = wb.create_sheet(HEADERS_SHEET)
    ws_h.sheet_state = "hidden"
    for i, h in enumerate(raw_headers, start=1):
        ws_h.cell(row=i, column=1, value=h)

    # Rebuild column_mapping sheet
    ws_m = wb[MAP_SHEET]
    ws_m.delete_rows(1, ws_m.max_row)

    headers_row = ["raw_column", "internal_name", "description", "required", "notes", "status"]
    for ci, h in enumerate(headers_row, start=1):
        cell = ws_m.cell(row=1, column=ci, value=h)
        cell.font = Font(bold=True)

    for ri, r in enumerate(results, start=2):
        fill = r.pop("_fill")
        for ci, key in enumerate(headers_row, start=1):
            cell = ws_m.cell(row=ri, column=ci, value=r.get(key, ""))
            cell.fill = fill

    # DataValidation dropdown on col A (raw_column), rows 2+
    dv = DataValidation(
        type="list",
        formula1=f"'{HEADERS_SHEET}'!$A$1:$A${len(raw_headers)}",
        allow_blank=True,
        showDropDown=False,
        showErrorMessage=False,   # allow any value (wildcards need free text)
    )
    dv.prompt = "Pick raw_data column header"
    dv.promptTitle = "Column Mapping"
    dv.sqref = f"A2:A{max(len(results) + 50, 200)}"
    ws_m.add_data_validation(dv)

    # Column widths + freeze
    for col, width in zip("ABCDEF", [30, 28, 45, 10, 55, 15]):
        ws_m.column_dimensions[col].width = width
    ws_m.freeze_panes = "A2"

    # Auto-filter
    ws_m.auto_filter.ref = f"A1:F{len(results) + 1}"

    wb.save(MASTER_PATH)
    print(f"\nSaved. Open column_mapping sheet.")
    if needs_review:
        print(f"  {needs_review} row(s) marked NEEDS REVIEW (red) -- pick correct header from dropdown.")
    if auto_matched:
        print(f"  {auto_matched} row(s) auto-matched (amber) -- verify these look correct.")


if __name__ == "__main__":
    reconcile()
