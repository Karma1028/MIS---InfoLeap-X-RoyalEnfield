"""One-shot builder for data/RE_MIS_Master.xlsx — the single editable
backend configuration file for the Royal Enfield MIS dashboard.
Run once to create; re-run to regenerate from scratch.
"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import openpyxl

HEADER_FILL = PatternFill("solid", fgColor="DEEAF1")
HEADER_FONT = Font(bold=True, color="1F3864", size=11)
HEADER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)
BODY_ALIGN = Alignment(horizontal="left", vertical="top", wrap_text=True)
THIN = Side(border_style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

MASTERFILE_PATH = "data/Enroute_Fourth Wave_Masterfile_Base_4010_AUG-MAY.xlsx"
OUTPUT_PATH = "data/RE_MIS_Master.xlsx"


def _style_header_row(ws, row=1):
    for cell in ws[row]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = HEADER_ALIGN
        cell.border = BORDER


def _auto_width(ws, min_w=12, max_w=50):
    for col_cells in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col_cells[0].column)
        for cell in col_cells:
            try:
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
            except Exception:
                pass
        ws.column_dimensions[col_letter].width = min(max_w, max(min_w, max_len + 2))


def _body_cell(ws, row, col, value):
    c = ws.cell(row=row, column=col, value=value)
    c.alignment = BODY_ALIGN
    c.border = BORDER
    return c


def build_readme(wb):
    ws = wb.create_sheet("README")
    ws.sheet_view.showGridLines = False
    title_cell = ws["A1"]
    title_cell.value = "RE MIS Master Configuration File — README"
    title_cell.font = Font(bold=True, size=14, color="1F3864")
    ws.row_dimensions[1].height = 24

    rows = [
        ("Sheet", "Purpose", "What to edit", "What NOT to edit"),
        ("column_mapping", "Maps raw Masterfile column names to internal app variable names",
         "Change raw_column if the Masterfile renames a survey column. Do NOT change internal_name.",
         "internal_name, required, description — these match app code"),
        ("segment_config", "Defines how Acceptor/Rejector/Cancelled segments are identified from grida column",
         "description only",
         "segment_name, grida_value — these must match app code constants"),
        ("netting_KBF", "Supernet/Net/Sub-net/Codelist/Codes taxonomy for Acceptor open-ended KBF question",
         "Add new rows for new netting codes. Edit Supernet/Net/Sub-net labels if taxonomy changes.",
         "Codes column — must match respondent-level code strings in raw data"),
        ("netting_Rejector", "Same taxonomy for Rejector Reasons for Rejection question",
         "Same as netting_KBF",
         "Same as netting_KBF"),
        ("netting_Cancelled", "Same taxonomy for Cancelled Reasons for Cancellation question",
         "Same as netting_KBF",
         "Same as netting_KBF"),
        ("month_order", "Display order for months in tables and charts",
         "This is auto-generated from actual data — normally no edit needed. Add sort_order for new months if auto-detection fails.",
         "month_label format must stay as \"MonthName'YYYY\" e.g. August'2024"),
    ]
    ws.append([])  # blank row 2
    for i, row_data in enumerate(rows):
        ws.append(list(row_data))
        r = ws.max_row
        if i == 0:
            _style_header_row(ws, r)
        else:
            for col_idx, val in enumerate(row_data, start=1):
                c = ws.cell(row=r, column=col_idx)
                c.alignment = BODY_ALIGN
                c.border = BORDER

    ws.freeze_panes = "A3"
    _auto_width(ws)
    return ws


def build_column_mapping(wb):
    ws = wb.create_sheet("column_mapping")
    headers = ["raw_column", "internal_name", "description", "required", "notes"]
    ws.append(headers)
    _style_header_row(ws)

    # All columns referenced in data_engine.py, data-engine logic, and _merge_reasons_codes
    mappings = [
        # Core segment/model derivation
        ("grida",         "grida",         "Segment gate: 1=Acceptor zone, 2=Rejector, 3=Cancelled",                         "YES", "Do not rename — segment logic depends on this"),
        ("acc",           "acc",           "Acceptor RE model code (1-14); null for non-Acceptors",                            "YES", "Used with value_map from datamap Sheet2"),
        ("seg",           "seg",           "Joint segment+model: 1-14=Acc, 15-28=Rej, 29-42=Can",                             "YES", "Used for Rejector/Cancelled re_model_code derivation"),
        ("aq3_po",        "aq3_po",        "RE model purchased (1-14); present on Acceptor-path respondents",                  "YES", "Used as fallback model code for reclassified Acceptors"),
        ("aq3",           "aq3",           "Brand/model purchased by Rejectors/Cancelled (1-124 scheme)",                      "YES", "Used for owned_brand_code on non-Acceptors"),
        # Temporal columns
        ("Month",         "Month",         "Numeric month of fielding (1=Jan, 12=Dec) — NOT SubmissionDate",                   "YES", "Source of truth for month_label derivation"),
        ("Year",          "Year",          "Numeric year of fielding (e.g. 2024)",                                             "YES", "Paired with Month for month_label"),
        ("SubmissionDate","SubmissionDate", "Survey submission timestamp — used for deduplication only",                        "YES", "Dedup key when merging monthly drops"),
        # Demographics
        ("dq1a",          "dq1a",          "Type of Buyer: 1=had 2W+Added, 2=had 2W+Replaced, 3=FTB family, 4=FTB no family", "YES", "Value labels in datamap Sheet2"),
        ("dq2b",          "dq2b",          "Single-select: which 2W was replaced (dq2 netting codebook scheme)",               "YES", "Used in Additional+Replaced Replaced portion"),
        ("dq3",           "dq3",           "Education level (1-7, see EDUCATION_DISPLAY_GROUPS in data_engine.py)",            "YES", "Value labels in datamap Sheet2"),
        ("dq4",           "dq4",           "Occupation (1-15, see OCCUPATION_DISPLAY_GROUPS in data_engine.py)",               "YES", "Value labels in datamap Sheet2"),
        ("dq6",           "dq6",           "Household income bracket",                                                         "YES", "Value labels in datamap Sheet2"),
        ("age_grp",       "age_grp",       "Age group (1=18-25, 2=26-35, 3=36-45, 4=46+) — derived/recoded",                  "YES", "Hardcoded value map in data_engine if not in datamap"),
        # Competitor / consideration
        ("aq1b",          "aq1b",          "Post-cancellation action (1=bought 2W, 2=bought car, 3=still searching, 4=dropped)", "YES", "Only meaningful for Cancelled segment"),
        ("aq2a",          "aq2a",          "Competitor CC range considered (2=150-249, 3=250-350, 4=351+)",                   "YES", "Competitor CC table"),
        ("aq5b",          "aq5b",          "Which RE model most seriously considered (Rejectors, codes 1-14)",                  "NO",  "Optional: only meaningful for Rejectors"),
        ("aq5c",          "aq5c",          "Brand resilience: what would you buy if your preferred unavailable (1-124)",        "NO",  "Optional: brand resilience table"),
        # Dynamic multi-select (auto-detected by prefix — editing raw_column not supported for these)
        ("dq2a_*",        "dq2a_*",        "Multi-select Additional vehicles (dq2a_1..dq2a_N, one col per vehicle code)",      "YES", "Auto-detected by prefix; all matching columns used"),
        ("aq5a_*",        "aq5a_*",        "Multi-select Brand Considered (aq5a_1..aq5a_124, 1=considered that brand)",        "YES", "Auto-detected by prefix; all matching columns used"),
        ("aq6_*",         "aq6_*",         "Test ride binary (aq6_1..aq6_14, 1=test rode that RE model)",                      "NO",  "Auto-detected by prefix"),
        # Netting code columns (respondent-level codes merged from masterfile)
        ("MQ2a+MQ2b_KBF",                "kbf_codes",        "Respondent's own KBF netting codes (concatenated 3-digit strings)", "YES", "Must match netting_KBF sheet's Codes column"),
        ("MQ3a+MQ3b_Rejecter",           "rejecter_codes",   "Respondent's own Rejection netting codes",                          "YES", "Must match netting_Rejector sheet's Codes column"),
        ("MQ3a+MQ3b_Booked and cancelled","cancelled_codes",  "Respondent's own Cancellation netting codes",                       "YES", "Must match netting_Cancelled sheet's Codes column"),
    ]

    for row_data in mappings:
        ws.append(list(row_data))
        r = ws.max_row
        for col_idx in range(1, 6):
            c = ws.cell(row=r, column=col_idx)
            c.alignment = BODY_ALIGN
            c.border = BORDER

    ws.freeze_panes = "A2"
    _auto_width(ws, min_w=16)
    return ws


def build_segment_config(wb):
    ws = wb.create_sheet("segment_config")
    headers = ["segment_name", "grida_value", "aq3_po_range_start", "aq3_po_range_end", "description"]
    ws.append(headers)
    _style_header_row(ws)

    rows = [
        ("Acceptor",  1, 1, 14, "Bought a Royal Enfield — identified by aq3_po between 1-14"),
        ("Rejector",  2, None, None, "Looked at RE but bought a different brand — grida==2"),
        ("Cancelled", 3, None, None, "Booked RE but cancelled before delivery — grida==3"),
    ]
    for row_data in rows:
        ws.append(list(row_data))
        r = ws.max_row
        for col_idx in range(1, 6):
            c = ws.cell(row=r, column=col_idx)
            c.alignment = BODY_ALIGN
            c.border = BORDER

    ws.freeze_panes = "A2"
    _auto_width(ws)
    return ws


def build_netting_sheet(wb, sheet_name, source_sheet_name):
    """Copy netting taxonomy from masterfile into a named sheet."""
    ws = wb.create_sheet(sheet_name)
    src_wb = openpyxl.load_workbook(MASTERFILE_PATH, read_only=True, data_only=True)
    src_ws = src_wb[source_sheet_name]

    for i, row in enumerate(src_ws.iter_rows(values_only=True)):
        ws.append(list(row))
        r = ws.max_row
        if i == 0:
            _style_header_row(ws, r)
        else:
            for col_idx, val in enumerate(row, start=1):
                c = ws.cell(row=r, column=col_idx)
                c.alignment = BODY_ALIGN
                c.border = BORDER

    src_wb.close()
    ws.freeze_panes = "A2"
    _auto_width(ws)
    note_cell = ws.cell(row=1, column=7, value="EDITABLE: Add/edit Supernet/Net/Sub-net/Codelist. Do NOT change Codes column values — they must match respondent code strings in raw data.")
    note_cell.font = Font(italic=True, color="7F7F7F", size=9)
    return ws


def build_month_order(wb):
    ws = wb.create_sheet("month_order")
    headers = ["month_label", "sort_order", "display_name", "notes"]
    ws.append(headers)
    _style_header_row(ws)

    # Known months from the current dataset (dynamic in app — this is reference)
    months = [
        ("August'2024",    1,  "August 2024",     ""),
        ("September'2024", 2,  "September 2024",  ""),
        ("October'2024",   3,  "October 2024",    ""),
        ("November'2024",  4,  "November 2024",   ""),
        ("December'2024",  5,  "December 2024",   ""),
        ("January'2025",   6,  "January 2025",    ""),
        ("February'2025",  7,  "February 2025",   ""),
        ("March'2025",     8,  "March 2025",      ""),
        ("April'2025",     9,  "April 2025",      ""),
        ("May'2025",       10, "May 2025",         "New data wave"),
    ]
    for row_data in months:
        ws.append(list(row_data))
        r = ws.max_row
        for col_idx in range(1, 5):
            c = ws.cell(row=r, column=col_idx)
            c.alignment = BODY_ALIGN
            c.border = BORDER

    note_cell = ws.cell(row=1, column=5, value="NOTE: App auto-detects months from raw data. This sheet is reference only. Add new months here for documentation when a new wave is loaded.")
    note_cell.font = Font(italic=True, color="7F7F7F", size=9)
    ws.freeze_panes = "A2"
    _auto_width(ws)
    return ws


def main():
    wb = openpyxl.Workbook()
    # Remove default sheet
    default = wb.active
    wb.remove(default)

    print("Building README sheet...")
    build_readme(wb)
    print("Building column_mapping sheet...")
    build_column_mapping(wb)
    print("Building segment_config sheet...")
    build_segment_config(wb)
    print("Building netting_KBF sheet...")
    build_netting_sheet(wb, "netting_KBF", "MQ2a+MQ2b_KBF")
    print("Building netting_Rejector sheet...")
    build_netting_sheet(wb, "netting_Rejector", "MQ3a+MQ3b_Rejecter")
    print("Building netting_Cancelled sheet...")
    build_netting_sheet(wb, "netting_Cancelled", "MQ3a+MQ3b_Booked and cancelled")
    print("Building month_order sheet...")
    build_month_order(wb)

    wb.save(OUTPUT_PATH)
    print(f"\nSaved: {OUTPUT_PATH}")
    print(f"Sheets: {wb.sheetnames}")


if __name__ == "__main__":
    main()
