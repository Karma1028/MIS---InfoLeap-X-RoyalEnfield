"""
Flatten and clean RE_MIS_Master.xlsx config sheets.
- datamap_value_labels: fill down variable name (remove merged-cell gaps)
- datamap_labels: fill down variable name
- column_mapping: strip instruction row 0
- display_groups: strip instruction row 0
Writes output to data/RE_MIS_Master.xlsx in-place (backs up first).
"""
import sys
import shutil
import pandas as pd
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

SRC = Path("data/RE_MIS_Master.xlsx")
BAK = Path("data/RE_MIS_Master_pre_clean_backup.xlsx")

sys.stdout.reconfigure(encoding='utf-8')

# ── backup ────────────────────────────────────────────────────────────────────
shutil.copy(SRC, BAK)
print(f"Backup: {BAK}")

# ── read all sheets as raw ────────────────────────────────────────────────────
xl = pd.ExcelFile(SRC)
sheets_raw = {}
for s in xl.sheet_names:
    sheets_raw[s] = xl.parse(s, header=None)

# ── helpers ───────────────────────────────────────────────────────────────────

def strip_instruction_row(df_raw, expected_cols):
    """Skip row 0 if it looks like an instruction (very long text in first cell)."""
    first = str(df_raw.iloc[0, 0])
    if len(first) > 40 or first.lower().startswith(('integer', 'internal', 'grouping', 'stable')):
        df = df_raw.iloc[1:].reset_index(drop=True)
    else:
        df = df_raw.copy()
    df.columns = expected_cols[:len(df.columns)]
    return df


def flatten_merged(df_raw, variable_col=0, code_col=1, label_col=2):
    """
    datamap_value_labels / datamap_labels style:
    Row pattern: variable in col0 (then blank), code in col1, label in col2.
    Fill down variable name.
    """
    # Find header row (where col0 has 'Variable' or first data row after title)
    start = 0
    for i, row in df_raw.iterrows():
        val = str(row[variable_col]).strip().lower()
        if val not in ('nan', 'variable values', 'variable information', 'value', 'variable', ''):
            start = i
            break

    df = df_raw.iloc[start:].reset_index(drop=True).copy()
    df = df[[variable_col, code_col, label_col]].copy()
    df.columns = ['variable', 'code', 'label']

    # Fill down variable
    df['variable'] = df['variable'].replace('', pd.NA).ffill()
    df = df.dropna(subset=['code'])
    df = df[df['variable'].notna()]
    df['variable'] = df['variable'].astype(str).str.strip()
    df['label'] = df['label'].astype(str).str.strip().replace('nan', '')
    return df.reset_index(drop=True)

# ── process each sheet ────────────────────────────────────────────────────────
cleaned = {}

# raw_data — keep as-is (too big to rewrite, just passthrough)
for s in ['raw_data', 'netting_KBF', 'netting_Rejector', 'netting_Cancelled', 'netting_aq3a_aq5', 'README']:
    cleaned[s] = xl.parse(s)

# datamap_value_labels
print("Cleaning datamap_value_labels...")
dvl = flatten_merged(sheets_raw['datamap_value_labels'])
print(f"  {len(dvl)} rows, {dvl['variable'].nunique()} unique variables")
cleaned['datamap_value_labels'] = dvl

# datamap_labels
print("Cleaning datamap_labels...")
dl_raw = sheets_raw['datamap_labels']
# cols: variable, position, label, measurement_level
dl_start = 0
for i, row in dl_raw.iterrows():
    val = str(row[0]).strip().lower()
    if val not in ('nan', 'variable information', 'variable', ''):
        dl_start = i
        break
dl = dl_raw.iloc[dl_start:].reset_index(drop=True)
dl.columns = ['variable', 'position', 'label', 'measurement_level'][:len(dl.columns)]
dl = dl[dl['variable'].notna() & (dl['variable'].astype(str).str.strip() != 'nan')]
print(f"  {len(dl)} rows")
cleaned['datamap_labels'] = dl

# column_mapping — strip instruction row 0
print("Cleaning column_mapping...")
cm = strip_instruction_row(sheets_raw['column_mapping'], ['category', 'raw_column', 'semantic_key', 'notes'])
cm = cm[cm['raw_column'].notna() & (cm['raw_column'].astype(str).str.strip() != 'nan')]
print(f"  {len(cm)} rows")
cleaned['column_mapping'] = cm

# display_groups — strip instruction row 0
print("Cleaning display_groups...")
dg = strip_instruction_row(sheets_raw['display_groups'], ['variable', 'code', 'display_label', 'notes'])
dg = dg[dg['variable'].notna() & (dg['variable'].astype(str).str.strip() != 'nan')]
print(f"  {len(dg)} rows")
cleaned['display_groups'] = dg

# model_config — strip instruction rows (row 0 header + trailing note rows)
print("Cleaning model_config...")
mc = xl.parse('model_config')
mc = mc[pd.to_numeric(mc['model_code'], errors='coerce').notna()].reset_index(drop=True)
print(f"  {len(mc)} rows")
cleaned['model_config'] = mc

# segment_config — keep as-is
cleaned['segment_config'] = xl.parse('segment_config')

# month_order — keep as-is
cleaned['month_order'] = xl.parse('month_order')

# ── write out with formatting ─────────────────────────────────────────────────
SHEET_ORDER = [
    'README', 'model_config', 'column_mapping', 'display_groups',
    'segment_config', 'month_order',
    'netting_KBF', 'netting_Rejector', 'netting_Cancelled', 'netting_aq3a_aq5',
    'raw_data', 'datamap_labels', 'datamap_value_labels',
]

print("\nWriting cleaned file...")
with pd.ExcelWriter(SRC, engine='openpyxl') as writer:
    for sheet in SHEET_ORDER:
        if sheet in cleaned:
            cleaned[sheet].to_excel(writer, sheet_name=sheet, index=False)

# ── apply header formatting ───────────────────────────────────────────────────
wb = load_workbook(SRC)

HEADER_FILL = PatternFill("solid", fgColor="1A1A2E")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
ALT_FILL    = PatternFill("solid", fgColor="F5F5F0")

for ws in wb.worksheets:
    if ws.title == 'raw_data':
        continue  # skip raw_data — too large

    # Header row
    for cell in ws[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=False)

    # Auto-width + alternating rows
    for i, row in enumerate(ws.iter_rows(min_row=2), start=2):
        for cell in row:
            cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=False)
            if i % 2 == 0:
                cell.fill = ALT_FILL

    # Column widths
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                max_len = max(max_len, len(str(cell.value or '')))
            except Exception:
                pass
        ws.column_dimensions[col_letter].width = min(max_len + 4, 60)

    # Freeze header
    ws.freeze_panes = "A2"

wb.save(SRC)
print(f"Done. Saved: {SRC}")
print(f"Backup at: {BAK}")
print("\nTo restore: copy RE_MIS_Master_pre_clean_backup.xlsx -> RE_MIS_Master.xlsx")
